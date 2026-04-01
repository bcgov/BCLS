"""
create_excel.py — Build the master Excel workbook from CSV files
================================================================
Run this ONCE to create data/BCLS_Dashboard_Data.xlsx from the
existing CSV files.  After that, edit the Excel file directly and
run excel_to_csv.py (or bundle.py) to regenerate the CSVs and
the standalone dashboard.

This file is generic: the CONFIG sheet drives sector-specific
settings, so the same Excel structure works for any sector
(e.g. AI & Quantum Technologies) just by updating CONFIG.

Usage (run from the BCLS root folder):
    python code/create_excel.py

Output:
    data/BCLS_Dashboard_Data.xlsx
"""

import csv
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT       = os.path.dirname(SCRIPT_DIR)
DATA_DIR   = os.path.join(ROOT, 'data')
OUTPUT_XL  = os.path.join(DATA_DIR, 'BCLS_Dashboard_Data.xlsx')

# ── Sheet → CSV mapping ───────────────────────────────────────────────────
# Sheet names are generic (sector-agnostic).
# Descriptions are written into each sheet's header row as a comment.
SHEETS = [
    # (sheet_name, csv_relative_path, description)
    ('Subsectors',           '00_reference/subsector_naics_lookup.csv',
     'Subsector definitions: IDs, NAICS codes, data availability, and tab settings. '
     'Each row is one subsector. show_as_tab=TRUE means it gets its own dashboard tab.'),

    ('Sector_Definition',    '00_reference/sector_definition_notes.csv',
     'Sector definition text shown at the top of the dashboard. '
     'Typically 1-3 rows: DEF-001 (main text), caveats, exclusions.'),

    ('Sector_Metrics',       '10_official_sector/official_sector_metrics.csv',
     'Core quantitative metrics (employment, GDP, revenue, etc.) by year, subsector, and province. '
     'subsector_id=ALL means the whole sector. Use metric_id codes consistently.'),

    ('Province_Compare',     '10_official_sector/official_sector_province_compare.csv',
     'Same metrics as Sector_Metrics but for multiple provinces, enabling BC vs Canada comparisons.'),

    ('Region_Shares',        '10_official_sector/official_sector_region_shares.csv',
     'Regional breakdown of key metrics within BC (Metro Vancouver, Victoria, etc.).'),

    ('Business_Counts',      '20_public_metrics/public_business_counts.csv',
     'Registered business counts by year, subsector, and firm size band.'),

    ('GDP_Metrics',          '20_public_metrics/public_gdp_metrics.csv',
     'GDP at basic prices by subsector and year (in millions CAD).'),

    ('IO_Multipliers',       '20_public_metrics/io_multipliers.csv',
     'Input-output multipliers for economic impact analysis.'),

    ('Goods_Exports',        '20_public_metrics/public_goods_exports.csv',
     'Life sciences goods exports by subsector, destination country, and year (millions CAD).'),

    ('Research_Activity',    '20_public_metrics/public_clinical_trials.csv',
     'Research/clinical activity metrics (e.g. active clinical trials with a BC site).'),

    ('Funding_Projects',     '20_public_metrics/public_funding_projects.csv',
     'Public funding by program, year, and recipient subsector (millions CAD).'),

    ('Region_Context',       '20_public_metrics/region_context.csv',
     'Contextual data for regional analysis (population, infrastructure, etc.).'),

    ('Talent_Pipeline',      '20_public_metrics/postsecondary_pipeline.csv',
     'Post-secondary graduates in sector-aligned programs by institution and credential type.'),

    ('Strategy_Goals',       '30_strategy/lookwest_goals.csv',
     'Strategy goals with targets, baselines, current progress, and RAG status. '
     'related_subsectors uses pipe-separated IDs (e.g. PHARMA|MEDDEV) or ALL.'),

    ('Strategy_Initiatives', '30_strategy/lookwest_initiatives.csv',
     'Strategy initiatives with owner, status, progress, and milestone tracking.'),

    ('Risks_Gaps',           '30_strategy/risks_and_gaps.csv',
     'Risk register and data gap log with severity, likelihood, RAG status, and mitigation notes.'),

    ('Evidence',             '40_manual/evidence_library.csv',
     'Evidence library: research reports, policy documents, and data releases supporting the strategy.'),

    ('Manual_Overrides',     '40_manual/manual_metric_overrides.csv',
     'Ministry-verified overrides for specific KPI values. These take priority over calculated data.'),

    ('Anchor_Companies',     '40_manual/anchor_companies_optional.csv',
     'List of anchor companies (100+ employees) used for company tracking. Optional.'),

    ('Whats_New',            '40_manual/whats_new_optional.csv',
     'Recent updates and news items displayed in the What\'s New panel. Optional.'),
]

# ── CONFIG sheet content ──────────────────────────────────────────────────
SECTOR_SETTINGS = [
    # (parameter, value, description)
    ('sector_name',        'BC Life Sciences',
     'Display name for the sector — used in dashboard titles and labels'),
    ('province',           'British Columbia',
     'Full province / jurisdiction name'),
    ('province_abbr',      'BC',
     'Province abbreviation used in metric data (must match "province" column values)'),
    ('strategy_name',      'Look West Strategy',
     'Name of the strategy being tracked (e.g. "Look West Strategy", "BC Tech Roadmap")'),
    ('strategy_year',      '2030',
     'Target year for all strategy goals'),
    ('dashboard_title',    'B.C. Life Sciences Sector Dashboard',
     'Main title shown in the browser tab and dashboard header'),
    ('dashboard_subtitle', 'Look West Strategy \u00b7 2030 Target Monitoring',
     'Subtitle shown below the main title in the header'),
    ('last_updated',       '2024-11-01',
     'Date of most recent data update (YYYY-MM-DD format)'),
    ('internal_only',      'TRUE',
     'Set TRUE to show the "Internal Working Document" badge in the header'),
]

CHART_SETTINGS = [
    # (chart_id, chart_name, show, tabs, description)
    ('emp_trend',       'Employment Trend',        'TRUE',  'ALL',
     'Time series of employment (FTE) for each subsector, 2019\u2013present with 2030 target line'),
    ('gdp_trend',       'GDP Contribution',        'TRUE',  'ALL',
     'GDP at basic prices trend by year for each subsector (millions CAD)'),
    ('biz_count',       'Business Count',          'TRUE',  'ALL',
     'Total registered businesses by year for each subsector (all size bands combined)'),
    ('exports',         'Goods Exports',           'TRUE',  'ALL',
     'Total life sciences goods exports by year for each subsector (millions CAD)'),
    ('regional',        'Regional Employment',     'TRUE',  'ALL',
     'FTE breakdown by BC region (Metro Vancouver, Victoria, etc.) \u2014 latest year'),
    ('province_compare','Province Comparison',     'TRUE',  'ALL',
     'BC vs other provinces employment comparison \u2014 latest year horizontal bar chart'),
    ('research_activity','Research Activity',      'TRUE',  'CRO',
     'Active research/clinical trials with a BC site \u2014 shown only on the CRO tab'),
    ('funding',         'Public Funding',          'FALSE', 'ALL',
     'Public funding by federal/provincial program \u2014 latest year (set TRUE to show)'),
    ('talent_pipeline', 'Talent Pipeline',         'FALSE', 'ALL',
     'Post-secondary life sciences graduates by year (set TRUE to show)'),
]

SUBSECTOR_COLORS = [
    # (subsector_id, color_hex, label)
    ('PHARMA',  '#1B4F8A', 'Pharma & Biopharma'),
    ('MEDDEV',  '#2E86AB', 'Medical Devices'),
    ('BIOTECH', '#8E44AD', 'Biotech R\u0026D'),
    ('CRO',     '#E67E22', 'Clinical Research \u0026 Labs'),
    ('OTHER',   '#718096', 'Other / Emerging'),
    ('ALL',     '#003366', 'Sector Total'),
]

# ── Style helpers ─────────────────────────────────────────────────────────
NAVY  = '00337A'
GOLD  = 'FCBA19'
LT_BG = 'EBF4FF'
HDR_BG= '003366'
ALT   = 'F5F8FC'
GREY  = '718096'
WHITE = 'FFFFFF'
SECTION_BG = 'EBF2FA'

def bold(size=11, color='000000', italic=False):
    return Font(name='Calibri', size=size, bold=True, color=color, italic=italic)

def reg(size=11, color='000000', italic=False):
    return Font(name='Calibri', size=size, color=color, italic=italic)

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=False)

def left(wrap=False):
    return Alignment(horizontal='left', vertical='top', wrap_text=wrap)

thin = Border(
    left=Side(style='thin', color='D0D7E0'),
    right=Side(style='thin', color='D0D7E0'),
    top=Side(style='thin', color='D0D7E0'),
    bottom=Side(style='thin', color='D0D7E0'),
)
thick_bottom = Border(bottom=Side(style='medium', color='003366'))

def style_header_row(ws, row, cols, bg=HDR_BG, fg=WHITE):
    """Apply header styling to a row."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = bold(10, fg)
        cell.fill      = fill(bg)
        cell.alignment = left()
        cell.border    = thin

def style_data_row(ws, row, cols, alt=False):
    """Apply alternating data row styling."""
    bg = ALT if alt else WHITE
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = reg(10)
        cell.fill      = fill(bg)
        cell.alignment = left(wrap=True)
        cell.border    = thin

def set_col_widths(ws, widths):
    """Set column widths."""
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def freeze(ws, cell='A2'):
    ws.freeze_panes = cell

# ── Read CSV ──────────────────────────────────────────────────────────────
def read_csv(path):
    if not os.path.exists(path):
        return [], []
    with open(path, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        rows = [dict(r) for r in reader]
    return headers, rows

# ── Build CONFIG sheet ────────────────────────────────────────────────────
def build_config_sheet(wb):
    ws = wb.create_sheet('CONFIG', 0)  # first sheet
    ws.sheet_view.showGridLines = False

    # ── Title banner ──────────────────────────────────────────────────────
    ws.merge_cells('A1:E1')
    t = ws['A1']
    t.value     = 'Dashboard Configuration'
    t.font      = Font(name='Calibri', size=14, bold=True, color=WHITE)
    t.fill      = fill(NAVY)
    t.alignment = center()
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:E2')
    sub = ws['A2']
    sub.value     = ('Sector-agnostic settings. Update this sheet to adapt the dashboard '
                     'to any sector. Run  python code/bundle.py  after saving.')
    sub.font      = reg(10, GREY, italic=True)
    sub.fill      = fill(SECTION_BG)
    sub.alignment = left(wrap=True)
    ws.row_dimensions[2].height = 30

    current_row = 4

    # ── Section: Sector Settings ──────────────────────────────────────────
    ws.merge_cells(f'A{current_row}:E{current_row}')
    s = ws[f'A{current_row}']
    s.value     = 'SECTOR SETTINGS'
    s.font      = bold(11, NAVY)
    s.fill      = fill(SECTION_BG)
    s.alignment = left()
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    # Header
    for col, hdr in enumerate(['parameter', 'value', 'description'], 1):
        c = ws.cell(row=current_row, column=col, value=hdr)
        c.font      = bold(10, WHITE)
        c.fill      = fill(HDR_BG)
        c.alignment = left()
        c.border    = thin
    current_row += 1

    for i, (param, val, desc) in enumerate(SECTOR_SETTINGS):
        style_data_row(ws, current_row, 3, alt=(i % 2 == 1))
        ws.cell(row=current_row, column=1, value=param).font = bold(10)
        ws.cell(row=current_row, column=2, value=val)
        ws.cell(row=current_row, column=3, value=desc)
        for col in range(1, 4):
            ws.cell(row=current_row, column=col).border = thin
        current_row += 1

    current_row += 1  # blank row

    # ── Section: Chart Settings ───────────────────────────────────────────
    ws.merge_cells(f'A{current_row}:E{current_row}')
    s = ws[f'A{current_row}']
    s.value     = "CHART SETTINGS  \u2014  set 'TRUE' or 'FALSE' to show or hide each chart"
    s.font      = bold(11, NAVY)
    s.fill      = fill(SECTION_BG)
    s.alignment = left()
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    # Header
    for col, hdr in enumerate(['chart_id', 'chart_name', 'show', 'tabs', 'description'], 1):
        c = ws.cell(row=current_row, column=col, value=hdr)
        c.font      = bold(10, WHITE)
        c.fill      = fill(HDR_BG)
        c.alignment = left()
        c.border    = thin
    current_row += 1

    for i, (cid, cname, show, tabs, desc) in enumerate(CHART_SETTINGS):
        style_data_row(ws, current_row, 5, alt=(i % 2 == 1))
        ws.cell(row=current_row, column=1, value=cid).font  = bold(10)
        ws.cell(row=current_row, column=2, value=cname)
        c_show = ws.cell(row=current_row, column=3, value=show)
        c_show.font = Font(name='Calibri', size=10, bold=True,
                           color='276749' if show == 'TRUE' else 'C53030')
        ws.cell(row=current_row, column=4, value=tabs)
        ws.cell(row=current_row, column=5, value=desc)
        for col in range(1, 6):
            ws.cell(row=current_row, column=col).border = thin
        current_row += 1

    current_row += 1

    # ── Section: Subsector Colors ─────────────────────────────────────────
    ws.merge_cells(f'A{current_row}:E{current_row}')
    s = ws[f'A{current_row}']
    s.value     = 'SUBSECTOR COLORS & LABELS'
    s.font      = bold(11, NAVY)
    s.fill      = fill(SECTION_BG)
    s.alignment = left()
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    for col, hdr in enumerate(['subsector_id', 'color_hex', 'label'], 1):
        c = ws.cell(row=current_row, column=col, value=hdr)
        c.font      = bold(10, WHITE)
        c.fill      = fill(HDR_BG)
        c.alignment = left()
        c.border    = thin
    current_row += 1

    for i, (sid, color, label) in enumerate(SUBSECTOR_COLORS):
        hex_clean = color.lstrip('#')
        style_data_row(ws, current_row, 3, alt=(i % 2 == 1))
        ws.cell(row=current_row, column=1, value=sid).font = bold(10)
        c_col = ws.cell(row=current_row, column=2, value=color)
        c_col.fill = fill(hex_clean)
        c_col.font = Font(name='Calibri', size=10,
                          color=WHITE if int(hex_clean[:2], 16) < 128 else '000000')
        ws.cell(row=current_row, column=3, value=label)
        for col in range(1, 4):
            ws.cell(row=current_row, column=col).border = thin
        current_row += 1

    # ── Column widths ─────────────────────────────────────────────────────
    set_col_widths(ws, [22, 36, 18, 14, 60])

    return ws

# ── Build a data sheet from CSV ───────────────────────────────────────────
def build_data_sheet(wb, sheet_name, csv_path, description):
    headers, rows = read_csv(csv_path)

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    if not headers:
        ws['A1'] = f'[No data found at {os.path.relpath(csv_path, ROOT)}]'
        ws['A1'].font = reg(10, 'C53030', italic=True)
        print(f'  ⚠  Sheet "{sheet_name}": CSV not found at {csv_path}')
        return ws

    # ── Row 1: sheet description ──────────────────────────────────────────
    ws.merge_cells(f'A1:{get_column_letter(max(len(headers), 1))}1')
    d = ws['A1']
    d.value     = description
    d.font      = reg(10, GREY, italic=True)
    d.fill      = fill(SECTION_BG)
    d.alignment = left(wrap=True)
    ws.row_dimensions[1].height = 30

    # ── Row 2: header row ─────────────────────────────────────────────────
    style_header_row(ws, 2, len(headers))
    for col, hdr in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=hdr)
        c.font      = bold(10, WHITE)
        c.fill      = fill(HDR_BG)
        c.alignment = left()

    # ── Data rows ─────────────────────────────────────────────────────────
    for i, row in enumerate(rows):
        r = i + 3
        style_data_row(ws, r, len(headers), alt=(i % 2 == 1))
        for col, hdr in enumerate(headers, 1):
            val = row.get(hdr, '') or ''
            # Try numeric conversion for display
            try:
                if isinstance(val, str) and '.' in val:
                    val = float(val)
                elif isinstance(val, str) and val.lstrip('-').isdigit():
                    val = int(val)
            except (ValueError, AttributeError):
                pass
            ws.cell(row=r, column=col, value=val)

    # ── Auto-width columns (capped at 50 chars) ───────────────────────────
    for col_idx, hdr in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(hdr))
        for row in rows:
            v = str(row.get(hdr, ''))
            max_len = max(max_len, min(len(v), 50))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 52)

    # ── Freeze header ─────────────────────────────────────────────────────
    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(len(headers))}2'

    print(f'  ✓  Sheet "{sheet_name}": {len(rows)} rows, {len(headers)} columns')
    return ws


# ── Main ──────────────────────────────────────────────────────────────────
def main():
    print('Building master Excel workbook…\n')

    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    # CONFIG sheet
    print('  ✓  CONFIG sheet')
    build_config_sheet(wb)

    print()

    # Data sheets
    for sheet_name, csv_rel, description in SHEETS:
        csv_path = os.path.join(DATA_DIR, csv_rel)
        build_data_sheet(wb, sheet_name, csv_path, description)

    # ── Save ──────────────────────────────────────────────────────────────
    wb.save(OUTPUT_XL)
    size_kb = os.path.getsize(OUTPUT_XL) / 1024
    print(f'\n✓  Saved: data/BCLS_Dashboard_Data.xlsx  ({size_kb:.0f} KB)')
    print()
    print('Next steps:')
    print('  • Open data/BCLS_Dashboard_Data.xlsx to review or edit data')
    print('  • When done editing, run:  python code/bundle.py')
    print('  • bundle.py automatically calls excel_to_csv.py first')
    print()
    print('To adapt for another sector:')
    print('  1. Rename the file (e.g. AI_Quantum_Dashboard_Data.xlsx)')
    print('  2. Update the CONFIG sheet (sector_name, strategy_name, etc.)')
    print('  3. Update Subsectors and Strategy_Goals sheets with the new sector\'s data')
    print('  4. Run bundle.py')


if __name__ == '__main__':
    main()

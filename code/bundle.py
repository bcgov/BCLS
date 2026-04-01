"""
bundle.py — Generate a self-contained standalone dashboard HTML
================================================================
Workflow for updating data:
    1. Edit data/BCLS_Dashboard_Data.xlsx
    2. Run:  python code/bundle.py
    3. Open  output/dashboard_standalone.html  in any browser
    4. Share output/dashboard_standalone.html with colleagues

What this script does:
    • Reads BCLS_Dashboard_Data.xlsx directly (no intermediate CSV files)
    • Parses the CONFIG sheet and all data sheets
    • Embeds data + the BC Gov logo as inline base64 in dashboard.html
    • Writes output/dashboard_standalone.html — no server needed
"""

import base64
import json
import os
import sys
from datetime import date

try:
    from openpyxl import load_workbook
except ImportError:
    print('ERROR: openpyxl not installed. Run:')
    print('       pip install openpyxl --break-system-packages')
    sys.exit(1)

# ── Paths ─────────────────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
ROOT        = os.path.dirname(SCRIPT_DIR)
DATA_DIR    = os.path.join(ROOT, 'data')
TEMPLATE    = os.path.join(SCRIPT_DIR, 'dashboard.html')
OUTPUT_DIR  = os.path.join(ROOT, 'output')
OUTPUT_FILE = os.path.join(OUTPUT_DIR, 'dashboard_standalone.html')
LOGO_PATH   = os.path.join(ROOT, 'docs', 'logo', 'BCID_H_RGB_rev.png')

# Maps Excel sheet name → STATE.data key used in the dashboard JS
SHEET_MAP = {
    'Subsectors':           'subsectors',
    'Sector_Definition':    'sectorDef',
    'Sector_Metrics':       'officialMetrics',
    'Region_Shares':        'regionShares',
    'Province_Compare':     'provinceCompare',
    'Business_Counts':      'bizCounts',
    'GDP_Metrics':          'gdpMetrics',
    'IO_Multipliers':       'ioMultipliers',
    'Goods_Exports':        'goodsExports',
    'Research_Activity':    'clinicalTrials',
    'Funding_Projects':     'fundingProjects',
    'Region_Context':       'regionContext',
    'Talent_Pipeline':      'psiPipeline',
    'Strategy_Goals':       'lwGoals',
    'Strategy_Initiatives': 'lwInitiatives',
    'Risks_Gaps':           'risksGaps',
    'Evidence':             'evidenceLib',
    'Manual_Overrides':     'manualOverrides',
    'Anchor_Companies':     'anchorCompanies',
    'Whats_New':            'whatsNew',
}


# ── Excel helpers ──────────────────────────────────────────────────────────
def cell_str(v):
    """Return a cell value as a clean string."""
    if v is None:
        return ''
    if isinstance(v, bool):
        return 'TRUE' if v else 'FALSE'
    return str(v).strip()


def parse_config_sheet(ws):
    """Parse the CONFIG sheet into sector/charts/colors dicts."""
    cfg = {'sector': {}, 'charts': {}, 'chart_tabs': {}, 'colors': {}, 'labels': {}}
    section   = None
    skip_next = False

    for row in ws.iter_rows():
        vals  = [cell_str(c.value) for c in row]
        first = vals[0].upper() if vals else ''

        if 'SECTOR SETTINGS'  in first: section='sector';  skip_next=False; continue
        if 'CHART SETTINGS'   in first: section='charts';  skip_next=True;  continue
        if 'SUBSECTOR COLORS' in first: section='colors';  skip_next=True;  continue
        if skip_next: skip_next=False; continue
        if not vals[0]: continue

        if section == 'sector':
            key, val = vals[0], (vals[1] if len(vals) > 1 else '')
            if key and not key.startswith('#') and key.lower() != 'key':
                cfg['sector'][key] = val

        elif section == 'charts':
            cid  = vals[0]
            show = (vals[2].upper() if len(vals) > 2 else 'TRUE') == 'TRUE'
            tabs = [t.strip() for t in (vals[3] if len(vals) > 3 else 'ALL').split(',') if t.strip()]
            if cid and cid != 'chart_id':
                cfg['charts'][cid]    = show
                cfg['chart_tabs'][cid] = tabs

        elif section == 'colors':
            sid   = vals[0]
            color = vals[1] if len(vals) > 1 else ''
            label = vals[2] if len(vals) > 2 else ''
            if sid and sid != 'subsector_id':
                cfg['colors'][sid] = color
                cfg['labels'][sid] = label

    return cfg


def parse_data_sheet(ws):
    """
    Parse a data sheet:
      Row 1 = description  (skip)
      Row 2 = column headers
      Row 3+ = data
    Returns a list of dicts with string values.
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [cell_str(h) for h in rows[1]]
    result  = []
    for row in rows[2:]:
        obj      = {}
        has_data = False
        for i, hdr in enumerate(headers):
            if not hdr:
                continue
            v = row[i] if i < len(row) else None
            obj[hdr] = cell_str(v)
            if obj[hdr]:
                has_data = True
        if has_data:
            result.append(obj)
    return result


def find_excel():
    """Locate the dashboard Excel workbook."""
    preferred = os.path.join(DATA_DIR, 'BCLS_Dashboard_Data.xlsx')
    if os.path.exists(preferred):
        return preferred
    for f in os.listdir(DATA_DIR):
        if f.endswith('.xlsx') and 'Dashboard_Data' in f:
            return os.path.join(DATA_DIR, f)
    return None


# ── Main ───────────────────────────────────────────────────────────────────
def bundle():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if not os.path.exists(TEMPLATE):
        print(f'ERROR: Template not found at {TEMPLATE}')
        sys.exit(1)

    # ── Step 1: Find and read Excel ───────────────────────────────────────
    xl_path = find_excel()
    if not xl_path:
        print('ERROR: No Dashboard_Data.xlsx found in data/')
        print('       Create one with code/create_excel.py or place your file there.')
        sys.exit(1)

    print('Step 1: Reading Excel workbook')
    print('─' * 50)
    print(f'  Source: {os.path.relpath(xl_path, ROOT)}')

    wb          = load_workbook(xl_path, data_only=True)
    inline_data = {}
    total_rows  = 0

    # CONFIG sheet
    if 'CONFIG' in wb.sheetnames:
        inline_data['dashboardConfig'] = parse_config_sheet(wb['CONFIG'])
        print(f'  ✓  CONFIG  →  dashboardConfig')
    else:
        print('  ⚠  CONFIG sheet not found')

    # Data sheets
    for sheet_name, data_key in SHEET_MAP.items():
        if sheet_name in wb.sheetnames:
            rows = parse_data_sheet(wb[sheet_name])
            inline_data[data_key] = rows
            total_rows += len(rows)
            print(f'  ✓  {sheet_name:<25} →  {data_key:<22} {len(rows):>4} rows')
        else:
            print(f'  ⚠  Sheet not found (skipped): {sheet_name}')

    print(f'\n  Total: {total_rows} rows across {len(inline_data)} datasets')
    print('─' * 50)
    print()

    # ── Step 2: Build inline data block ──────────────────────────────────
    today       = date.today().isoformat()
    inline_json = json.dumps(inline_data, ensure_ascii=False, separators=(',', ':'))
    inject_block = (
        f'\n<!-- == INLINE DATA (generated by bundle.py on {today}) == -->\n'
        f'<script>\n'
        f'/* Auto-generated — do not edit manually.\n'
        f'   To update: edit BCLS_Dashboard_Data.xlsx, then run: python code/bundle.py */\n'
        f'window.BCLS_INLINE_DATA = {inline_json};\n'
        f'</script>\n'
        f'<!-- ========================================================= -->\n'
    )

    # ── Step 3: Inject into template ─────────────────────────────────────
    with open(TEMPLATE, encoding='utf-8') as f:
        html = f.read()

    if '</head>' not in html:
        print('ERROR: </head> tag not found in template.')
        sys.exit(1)

    output_html = html.replace('</head>', inject_block + '</head>', 1)

    # ── Step 4: Embed BC Gov logo as base64 ──────────────────────────────
    if os.path.exists(LOGO_PATH):
        with open(LOGO_PATH, 'rb') as f:
            logo_b64 = base64.b64encode(f.read()).decode()
        output_html = output_html.replace(
            'src="../docs/logo/BCID_H_RGB_rev.png"',
            f'src="data:image/png;base64,{logo_b64}"'
        )
        print(f'  ✓  BC Gov logo embedded (base64)')
    else:
        print(f'  ⚠  Logo not found at docs/logo/BCID_H_RGB_rev.png')

    # Update page title
    output_html = output_html.replace('<title>', f'<title>[Standalone · {today}] ')

    # ── Step 5: Write output ──────────────────────────────────────────────
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(output_html)

    size_kb = os.path.getsize(OUTPUT_FILE) / 1024
    print(f'\nStep 2: Standalone file built')
    print('─' * 50)
    print(f'  ✓  output/dashboard_standalone.html  ({size_kb:.0f} KB)')
    print(f'     Open this file in any browser — no server required.\n')
    print('─' * 50)
    print()
    print('Done! To share the dashboard, send  output/dashboard_standalone.html  to colleagues.')
    print('      To refresh data in the browser, click "↺ Refresh Data" and pick the Excel file.')


if __name__ == '__main__':
    bundle()

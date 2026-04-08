#!/usr/bin/env python3
"""Create a lightweight Life Sciences workbook aligned with the dashboard data contract."""

from pathlib import Path
from openpyxl import Workbook

OUT_PATH = Path(__file__).with_name("Life_Sciences_light.xlsx")


def add_sheet(ws, headers, rows):
    ws.append(headers)
    for r in rows:
        ws.append(r)


def chart_catalog_rows():
    return [
        ["emp_trend", "ALL", "TRUE", 1, "Employment Trend (FTE)", "{subsector} · 2019–2024 with target", "Official metrics + strategy targets", "DATA_EMP_TREND", ""],
        ["gdp_trend", "ALL", "TRUE", 2, "GDP Contribution ($M CAD)", "{subsector} · GDP at basic prices", "Statistics Canada CSNA Table 36-10-0402", "DATA_GDP_TREND", ""],
        ["biz_count", "ALL", "TRUE", 3, "Business Count", "{subsector} · Total registered businesses", "Statistics Canada Business Register", "DATA_BIZ_COUNT", ""],
        ["exports", "ALL", "TRUE", 4, "Goods Exports ($M CAD)", "{subsector} · Total by year", "Statistics Canada Trade Data Online", "DATA_EXPORTS", ""],
        ["regional", "ALL", "TRUE", 5, "Employment by Region (2024)", "B.C. regions · FTE share", "BC Stats proxy business location data", "DATA_REGIONAL", ""],
        ["province_compare", "ALL", "TRUE", 6, "Province Comparison (2024)", "Life sciences FTEs · BC vs other provinces", "Statistics Canada LFS", "DATA_PROVINCE_COMPARE", ""],
        ["research_activity", "ALL", "TRUE", 7, "Research Activity (Active Trials)", "Active clinical trials with B.C. site", "ClinicalTrials.gov / Health Canada CTDB", "DATA_RESEARCH_ACTIVITY", ""],
        ["funding", "ALL", "FALSE", 8, "Public Funding by Program ($M CAD)", "Federal + provincial programs · 2024", "ISED, CIHR, NSERC, BC MTIT", "DATA_FUNDING", "Set TRUE when data is ready"],
        ["talent_pipeline", "ALL", "FALSE", 9, "Talent Pipeline (Graduates)", "Post-secondary life sciences graduates · B.C.", "BCCAT / HEABC graduate data", "DATA_TALENT_PIPELINE", "Set TRUE when data is ready"],
    ]


def create_workbook(out_path: Path = OUT_PATH) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    add_sheet(
        wb.create_sheet("CONFIG"),
        ["key", "value", "description"],
        [
            ["sector_name", "BC Life Sciences", "Display name for the sector"],
            ["province", "British Columbia", "Full province name"],
            ["province_abbr", "BC", "Province abbreviation"],
            ["strategy_name", "Look West Strategy", "Strategy being tracked"],
            ["strategy_year", 2030, "Target year"],
            ["dashboard_title", "B.C. Life Sciences Sector Dashboard", "Main title"],
            ["dashboard_subtitle", "Look West Strategy - 2030 Target Monitoring", "Subtitle"],
            ["last_updated", "2026-04-03", "Date of most recent update"],
            ["internal_only", "TRUE", "Internal document flag"],
        ],
    )

    add_sheet(
        wb.create_sheet("SECTOR_STRUCTURE"),
        [
            "record_type", "sector_id", "subsector_id", "subsector_name", "naics_4digit", "naics_5digit",
            "naics_description", "tab_order", "show_as_tab", "dashboard_color", "has_reliable_data",
            "data_availability_note", "heading", "body_text", "caveat", "last_updated", "source_tier", "confidence_tag"
        ],
        [
            ["sector", "life_sciences", "ALL", "BC Life Sciences", "", "", "", 0, "TRUE", "#003366", "TRUE", "", "", "", "", "2026-04-03", "official-curated", "high"],
            ["subsector", "life_sciences", "PHARMA", "Pharma & Biopharma", 3254, 32541, "Pharmaceutical and medicine manufacturing", 1, "TRUE", "#1B4F8A", "TRUE", "Good data", "", "", "", "", "", ""],
            ["subsector", "life_sciences", "MEDDEV", "Medical Devices", 3391, 33910, "Medical equipment and supplies manufacturing", 2, "TRUE", "#2E86AB", "TRUE", "Good data", "", "", "", "", "", ""],
            ["subsector", "life_sciences", "BIOTECH", "Biotech R&D", 5417, 54171, "Research and development in life sciences", 3, "TRUE", "#8E44AD", "PARTIAL", "Partial data", "", "", "", "", "", ""],
            ["subsector", "life_sciences", "CRO", "Clinical Research & Labs", 6215, 62151, "Medical and diagnostic laboratories", 4, "TRUE", "#E67E22", "PARTIAL", "Partial data", "", "", "", "", "", ""],
            ["subsector", "life_sciences", "OTHER", "Other / Emerging", "", "", "Other life sciences activities", 5, "TRUE", "#718096", "FALSE", "Data gap", "", "", "", "", "", ""],
            ["definition", "life_sciences", "", "", "", "", "", "", "", "", "", "", "What is the BC Life Sciences Sector?", "The British Columbia life sciences sector includes pharmaceuticals, medical devices, biotech R&D, and clinical research services.", "", "2026-04-03", "official-curated", "high"],
        ],
    )

    add_sheet(
        wb.create_sheet("KEY_KPIS"),
        ["metric_id", "subsector_id", "province", "period", "value", "baseline_period", "baseline_value", "unit", "source_tier", "confidence_tag"],
        [
            ["EMP-TOTAL-BC", "ALL", "BC", 2024, 22100, 2019, 18500, "FTE", "official-curated", "high"],
            ["GDP-TOTAL-BC", "ALL", "BC", 2024, 4800, 2019, 3900, "millions_cad", "official-curated", "high"],
            ["REV-TOTAL-BC", "ALL", "BC", 2024, 12500, 2019, 9800, "millions_cad", "official-curated", "medium"],
            ["BIZ-TOTAL-BC", "ALL", "BC", 2024, 1150, 2019, 980, "count", "official-curated", "high"],
            ["EXP-GOODS-BC", "ALL", "BC", 2024, 890, 2019, 650, "millions_cad", "official-curated", "medium"],
            ["FUND-PUBLIC-BC", "ALL", "BC", 2024, 125, 2019, 80, "millions_cad", "public-direct", "medium"],
        ],
    )

    add_sheet(
        wb.create_sheet("CHART_CATALOG"),
        ["chart_id", "subsector_id", "is_visible", "display_order", "title", "subtitle", "caveat", "data_sheet", "notes"],
        chart_catalog_rows(),
    )

    chart_headers = ["metric_id", "subsector_id", "province", "region_name", "comparator", "period", "value", "value_2", "unit", "source_tier", "notes"]

    add_sheet(
        wb.create_sheet("DATA_EMP_TREND"),
        chart_headers,
        [],
    )
    add_sheet(
        wb.create_sheet("DATA_GDP_TREND"),
        chart_headers,
        [
            ["GDP-TOTAL-BC", "ALL", "BC", "", "", 2019, 3900, "", "millions_cad", "official-curated", ""],
            ["GDP-TOTAL-BC", "ALL", "BC", "", "", 2020, 3650, "", "millions_cad", "official-curated", ""],
            ["GDP-TOTAL-BC", "ALL", "BC", "", "", 2021, 4100, "", "millions_cad", "official-curated", ""],
            ["GDP-TOTAL-BC", "ALL", "BC", "", "", 2022, 4450, "", "millions_cad", "official-curated", ""],
            ["GDP-TOTAL-BC", "ALL", "BC", "", "", 2023, 4620, "", "millions_cad", "official-curated", ""],
            ["GDP-TOTAL-BC", "ALL", "BC", "", "", 2024, 4800, "", "millions_cad", "official-curated", ""],
        ],
    )
    add_sheet(
        wb.create_sheet("DATA_BIZ_COUNT"),
        chart_headers,
        [
            ["BIZ-TOTAL-BC", "ALL", "BC", "", "All sizes", 2019, 980, "", "count", "official-curated", ""],
            ["BIZ-TOTAL-BC", "ALL", "BC", "", "All sizes", 2020, 995, "", "count", "official-curated", ""],
            ["BIZ-TOTAL-BC", "ALL", "BC", "", "All sizes", 2021, 1020, "", "count", "official-curated", ""],
            ["BIZ-TOTAL-BC", "ALL", "BC", "", "All sizes", 2022, 1085, "", "count", "official-curated", ""],
            ["BIZ-TOTAL-BC", "ALL", "BC", "", "All sizes", 2023, 1120, "", "count", "official-curated", ""],
            ["BIZ-TOTAL-BC", "ALL", "BC", "", "All sizes", 2024, 1150, "", "count", "official-curated", ""],
        ],
    )
    add_sheet(
        wb.create_sheet("DATA_EXPORTS"),
        chart_headers,
        [
            ["EXP-GOODS-BC", "ALL", "BC", "", "All countries", 2019, 650, "", "millions_cad", "official-curated", ""],
            ["EXP-GOODS-BC", "ALL", "BC", "", "All countries", 2020, 620, "", "millions_cad", "official-curated", ""],
            ["EXP-GOODS-BC", "ALL", "BC", "", "All countries", 2021, 710, "", "millions_cad", "official-curated", ""],
            ["EXP-GOODS-BC", "ALL", "BC", "", "All countries", 2022, 780, "", "millions_cad", "official-curated", ""],
            ["EXP-GOODS-BC", "ALL", "BC", "", "All countries", 2023, 840, "", "millions_cad", "official-curated", ""],
            ["EXP-GOODS-BC", "ALL", "BC", "", "All countries", 2024, 890, "", "millions_cad", "official-curated", ""],
        ],
    )
    add_sheet(
        wb.create_sheet("DATA_REGIONAL"),
        chart_headers,
        [
            ["EMP-TOTAL-BC", "ALL", "BC", "Metro Vancouver", "", 2024, 16500, 74.7, "FTE", "official-curated", ""],
            ["EMP-TOTAL-BC", "ALL", "BC", "Victoria", "", 2024, 2800, 12.7, "FTE", "official-curated", ""],
            ["EMP-TOTAL-BC", "ALL", "BC", "Fraser Valley", "", 2024, 1600, 7.2, "FTE", "official-curated", ""],
        ],
    )
    add_sheet(
        wb.create_sheet("DATA_PROVINCE_COMPARE"),
        chart_headers,
        [
            ["EMP-TOTAL-BC", "ALL", "BC", "", "", 2024, 22100, "", "FTE", "official-curated", ""],
            ["EMP-TOTAL-BC", "ALL", "ON", "", "", 2024, 54200, "", "FTE", "official-curated", ""],
            ["EMP-TOTAL-BC", "ALL", "QC", "", "", 2024, 30100, "", "FTE", "official-curated", ""],
            ["EMP-TOTAL-BC", "ALL", "AB", "", "", 2024, 11800, "", "FTE", "official-curated", ""],
        ],
    )
    add_sheet(
        wb.create_sheet("DATA_RESEARCH_ACTIVITY"),
        chart_headers,
        [
            ["TRIALS-ACTIVE-BC", "ALL", "BC", "", "ALL|ALL", 2019, 280, "", "count", "public-direct", ""],
            ["TRIALS-ACTIVE-BC", "ALL", "BC", "", "ALL|ALL", 2020, 295, "", "count", "public-direct", ""],
            ["TRIALS-ACTIVE-BC", "ALL", "BC", "", "ALL|ALL", 2021, 330, "", "count", "public-direct", ""],
            ["TRIALS-ACTIVE-BC", "ALL", "BC", "", "ALL|ALL", 2022, 360, "", "count", "public-direct", ""],
            ["TRIALS-ACTIVE-BC", "ALL", "BC", "", "ALL|ALL", 2023, 390, "", "count", "public-direct", ""],
            ["TRIALS-ACTIVE-BC", "ALL", "BC", "", "ALL|ALL", 2024, 402, "", "count", "public-direct", ""],
        ],
    )
    add_sheet(
        wb.create_sheet("DATA_FUNDING"),
        chart_headers,
        [
            ["FUND-PUBLIC-BC", "ALL", "BC", "", "CIHR Operating Grants|Federal", 2024, 45.2, "", "millions_cad", "public-direct", ""],
            ["FUND-PUBLIC-BC", "ALL", "BC", "", "Strategic Innovation Fund|Federal", 2024, 38.0, "", "millions_cad", "public-direct", ""],
            ["FUND-PUBLIC-BC", "ALL", "BC", "", "BC Life Sciences Fund|Provincial", 2024, 41.8, "", "millions_cad", "public-direct", ""],
        ],
    )
    add_sheet(
        wb.create_sheet("DATA_TALENT_PIPELINE"),
        chart_headers,
        [
            ["GRADS-BC", "ALL", "BC", "", "ALL|ALL", 2019, 2850, "", "count", "official-curated", ""],
            ["GRADS-BC", "ALL", "BC", "", "ALL|ALL", 2020, 3000, "", "count", "official-curated", ""],
            ["GRADS-BC", "ALL", "BC", "", "ALL|ALL", 2021, 3200, "", "count", "official-curated", ""],
            ["GRADS-BC", "ALL", "BC", "", "ALL|ALL", 2022, 3350, "", "count", "official-curated", ""],
            ["GRADS-BC", "ALL", "BC", "", "ALL|ALL", 2023, 3450, "", "count", "official-curated", ""],
            ["GRADS-BC", "ALL", "BC", "", "ALL|ALL", 2024, 3620, "", "count", "official-curated", ""],
        ],
    )

    add_sheet(
        wb.create_sheet("STRATEGY_GOALS"),
        ["goal_id", "goal_name", "pillar", "rag_status", "progress_pct", "target_value", "target_unit", "target_year", "baseline_year", "baseline_value", "related_subsectors"],
        [
            ["LW-G01", "Grow sector employment to 30,000 FTE by 2030", "Workforce", "GREEN", 74, 30000, "FTE", 2030, 2019, 18500, "ALL"],
            ["LW-G07", "Grow active clinical trials to 600 by 2030", "Research", "GREEN", 67, 600, "trials", 2030, 2019, 280, "CRO"],
            ["LW-G08", "Increase life sciences graduates by 25%", "Workforce", "GREEN", 82, 3562, "graduates", 2030, 2019, 2850, "ALL"],
        ],
    )

    add_sheet(
        wb.create_sheet("INITIATIVE_TRACKER"),
        ["initiative_id", "initiative_name", "pillar", "rag_status", "status", "progress_pct", "owner_ministry", "next_milestone", "next_milestone_date", "last_review_date", "related_subsectors", "description"],
        [
            ["LW-I01", "BC Life Sciences Fund $100M allocation", "Investment", "GREEN", "Active", 85, "Ministry of Jobs", "Final disbursements", "2025-06-30", "2025-01-15", "ALL", "Deploy provincial fund to support sector growth."],
            ["LW-I04", "BC Clinical Trials Network Expansion", "Research", "GREEN", "Active", 78, "Ministry of Health", "Network expansion complete", "2025-03-31", "2025-01-20", "CRO", "Expand trial capacity across BC sites."],
        ],
    )

    add_sheet(
        wb.create_sheet("EVIDENCE_LIBRARY"),
        ["type", "title", "summary", "author", "publication_date", "confidence_tag", "source_tier", "related_goal_id", "related_initiative_id", "related_subsectors", "link"],
        [
            ["Research Report", "BC Life Sciences Sector Profile 2024", "Profile of employment, revenue and regional distribution.", "BC Stats", "2024-09-15", "high", "official-curated", "LW-G01|LW-G03", "", "ALL", "https://www2.gov.bc.ca/gov/content/data/statistics"],
        ],
    )

    add_sheet(
        wb.create_sheet("RISKS_GAPS"),
        ["risk_name", "category", "description", "severity", "likelihood", "rag_status", "owner", "mitigation", "last_reviewed", "related_subsectors"],
        [
            ["Talent shortage in specialized roles", "Workforce", "Competition for specialized life sciences talent may constrain growth.", "HIGH", "Likely", "AMBER", "Ministry of Post-Secondary", "Expand recruitment and training pathways.", "2025-01-15", "ALL"],
        ],
    )

    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    path = create_workbook()
    print(f"Workbook created: {path}")

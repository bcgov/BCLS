#!/usr/bin/env python3
"""
BC Life Sciences Sector Dashboard Data Source Generator
Creates Life_Sciences.xlsx with all required sheets and sample data
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# Color definitions
HEADER_COLOR = "FF003366"
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
ALTERNATING_COLORS = ["FFFFFFFF", "FFF0F6FF"]

def apply_formatting(ws, data, has_headers=True):
    """Apply professional formatting to worksheet"""
    # Apply header formatting
    if has_headers:
        for col in range(1, len(data[0]) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Apply alternating row colors and alignment
    start_row = 2 if has_headers else 1
    for row_idx, row in enumerate(data[1:] if has_headers else data, start=start_row):
        color_idx = (row_idx - start_row) % 2
        fill_color = ALTERNATING_COLORS[color_idx]

        for col in range(1, len(row) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # Auto-size columns with max width of 40
    for col in range(1, len(data[0]) + 1):
        column_letter = get_column_letter(col)
        max_length = 0

        for row in data:
            try:
                if len(str(row[col - 1])) > max_length:
                    max_length = len(str(row[col - 1]))
            except:
                pass

        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze top row
    if has_headers:
        ws.freeze_panes = "A2"

def create_excel():
    """Create the Life_Sciences.xlsx workbook with all sheets"""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: CONFIG
    ws = wb.create_sheet("CONFIG")
    config_data = [
        ["parameter", "value", "description"],
        ["sector_name", "BC Life Sciences", "Display name for the sector"],
        ["province", "British Columbia", "Full province name"],
        ["province_abbr", "BC", "Province abbreviation"],
        ["strategy_name", "Look West Strategy", "Strategy being tracked"],
        ["strategy_year", 2030, "Target year"],
        ["dashboard_title", "B.C. Life Sciences Sector Dashboard", "Main title"],
        ["dashboard_subtitle", "Look West Strategy · 2030 Target Monitoring", "Subtitle"],
        ["last_updated", "2025-01-01", "Date of most recent update"],
        ["internal_only", "TRUE", "Internal document flag"],
    ]
    for row in config_data:
        ws.append(row)
    apply_formatting(ws, config_data)

    # Sheet 2: Subsectors
    ws = wb.create_sheet("Subsectors")
    subsectors_data = [
        ["subsector_id", "subsector_name", "naics_4digit", "naics_5digit", "naics_description", "dashboard_color", "show_as_tab", "tab_order", "has_reliable_data", "data_availability_note"],
        ["PHARMA", "Pharma & Biopharma", 3254, 32541, "Pharmaceutical and medicine manufacturing", "#1B4F8A", "TRUE", 1, "TRUE", "Official BC Stats data available"],
        ["MEDDEV", "Medical Devices", 3391, 33911, "Medical equipment and supplies manufacturing", "#2E86AB", "TRUE", 2, "TRUE", "Official BC Stats data available"],
        ["BIOTECH", "Biotech R&D", 5417, 54171, "Research and development in life sciences", "#8E44AD", "TRUE", 3, "PARTIAL", "Some data estimated from national ratios"],
        ["CRO", "Clinical Research & Labs", 5415, 54151, "Computer systems design and related services", "#E67E22", "TRUE", 4, "PARTIAL", "CRO data estimated; clinical trials from clinicaltrials.gov"],
        ["OTHER", "Other / Emerging", 5413, 54131, "Architectural, engineering and related services", "#718096", "FALSE", 5, "FALSE", "Insufficient data for separate reporting"],
        ["ALL", "Sector Total", 54, "", "All life sciences subsectors combined", "#003366", "TRUE", 0, "TRUE", "Combined sector total"],
    ]
    for row in subsectors_data:
        ws.append(row)
    apply_formatting(ws, subsectors_data)

    # Sheet 3: Sector_Definition
    ws = wb.create_sheet("Sector_Definition")
    sector_def_data = [
        ["id", "category", "heading", "body_text", "caveat", "last_updated", "source_tier", "confidence_tag"],
        ["DEF-001", "definition", "What is the BC Life Sciences Sector?", "The British Columbia life sciences sector encompasses pharmaceutical manufacturing, medical device production, biotechnology research, and clinical research organizations. BC is home to over 1,100 life sciences businesses employing more than 20,000 people.", "This definition follows the BC Stats sector classification methodology. NAICS codes 3254, 3391, 5417 are the primary codes.", "2025-01-01", "official-curated", "high"],
        ["DEF-002", "scope", "Geographic Scope", "Data covers all of British Columbia unless otherwise noted. Regional breakdowns follow Statistics Canada's Economic Region definitions.", "Metro Vancouver represents approximately 75% of sector employment.", "2025-01-01", "official-curated", "high"],
        ["DEF-003", "exclusion", "Key Exclusions", "Retail pharmacies, hospitals, and medical clinics are excluded from sector counts as they are classified under healthcare services (NAICS 62).", "Exclusions follow BC Stats standard sector boundary definitions.", "2025-01-01", "official-curated", "high"],
    ]
    for row in sector_def_data:
        ws.append(row)
    apply_formatting(ws, sector_def_data)

    # Sheet 4: Sector_Metrics
    ws = wb.create_sheet("Sector_Metrics")
    sector_metrics_data = [
        ["metric_id", "subsector_id", "province", "year", "value", "baseline_year", "baseline_value", "unit", "source_tier", "confidence_tag"],
        ["EMP-TOTAL-BC", "ALL", "BC", 2019, 18500, 2019, 18500, "FTE", "official-curated", "high"],
        ["EMP-TOTAL-BC", "ALL", "BC", 2020, 17800, 2019, 18500, "FTE", "official-curated", "high"],
        ["EMP-TOTAL-BC", "ALL", "BC", 2021, 19200, 2019, 18500, "FTE", "official-curated", "high"],
        ["EMP-TOTAL-BC", "ALL", "BC", 2022, 20800, 2019, 18500, "FTE", "official-curated", "high"],
        ["EMP-TOTAL-BC", "ALL", "BC", 2023, 21500, 2019, 18500, "FTE", "official-curated", "high"],
        ["EMP-TOTAL-BC", "ALL", "BC", 2024, 22100, 2019, 18500, "FTE", "official-curated", "high"],
        ["GDP-TOTAL-BC", "ALL", "BC", 2024, 4800, 2019, 3900, "millions_cad", "official-curated", "high"],
        ["REV-TOTAL-BC", "ALL", "BC", 2024, 12500, 2019, 9800, "millions_cad", "official-curated", "medium"],
        ["BIZ-TOTAL-BC", "ALL", "BC", 2024, 1150, 2019, 980, "count", "official-curated", "high"],
        ["EXP-GOODS-BC", "ALL", "BC", 2024, 890, 2019, 650, "millions_cad", "official-curated", "medium"],
    ]
    for row in sector_metrics_data:
        ws.append(row)
    apply_formatting(ws, sector_metrics_data)

    # Sheet 5: Province_Compare
    ws = wb.create_sheet("Province_Compare")
    province_compare_data = [
        ["metric_id", "subsector_id", "province", "year", "value", "unit", "source_tier"],
        ["EMP-TOTAL-BC", "ALL", "BC", 2024, 22100, "FTE", "official-curated"],
        ["EMP-TOTAL-BC", "ALL", "ON", 2024, 54200, "FTE", "official-curated"],
        ["EMP-TOTAL-BC", "ALL", "QC", 2024, 31800, "FTE", "official-curated"],
        ["EMP-TOTAL-BC", "ALL", "AB", 2024, 12400, "FTE", "official-curated"],
        ["EMP-TOTAL-BC", "ALL", "Canada", 2024, 142000, "FTE", "official-curated"],
    ]
    for row in province_compare_data:
        ws.append(row)
    apply_formatting(ws, province_compare_data)

    # Sheet 6: Region_Shares
    ws = wb.create_sheet("Region_Shares")
    region_shares_data = [
        ["metric_id", "subsector_id", "region_name", "year", "value", "share_pct", "unit", "source_tier"],
        ["EMP-TOTAL-BC", "ALL", "Metro Vancouver", 2024, 16500, 74.7, "FTE", "official-curated"],
        ["EMP-TOTAL-BC", "ALL", "Victoria", 2024, 2800, 12.7, "FTE", "official-curated"],
        ["EMP-TOTAL-BC", "ALL", "Kelowna", 2024, 1100, 5.0, "FTE", "official-curated"],
        ["EMP-TOTAL-BC", "ALL", "Prince George", 2024, 600, 2.7, "FTE", "official-curated"],
        ["EMP-TOTAL-BC", "ALL", "Other BC", 2024, 1100, 5.0, "FTE", "official-curated"],
    ]
    for row in region_shares_data:
        ws.append(row)
    apply_formatting(ws, region_shares_data)

    # Sheet 7: Business_Counts
    ws = wb.create_sheet("Business_Counts")
    business_counts_data = [
        ["subsector_id", "year", "size_band", "count", "source_tier"],
        ["ALL", 2019, "All sizes", 980, "official-curated"],
        ["ALL", 2020, "All sizes", 995, "official-curated"],
        ["ALL", 2021, "All sizes", 1020, "official-curated"],
        ["ALL", 2022, "All sizes", 1085, "official-curated"],
        ["ALL", 2023, "All sizes", 1120, "official-curated"],
        ["ALL", 2024, "All sizes", 1150, "official-curated"],
    ]
    for row in business_counts_data:
        ws.append(row)
    apply_formatting(ws, business_counts_data)

    # Sheet 8: GDP_Metrics
    ws = wb.create_sheet("GDP_Metrics")
    gdp_metrics_data = [
        ["subsector_id", "year", "value_millions_cad", "source_tier"],
        ["ALL", 2019, 3900, "official-curated"],
        ["ALL", 2020, 3650, "official-curated"],
        ["ALL", 2021, 4100, "official-curated"],
        ["ALL", 2022, 4450, "official-curated"],
        ["ALL", 2023, 4620, "official-curated"],
        ["ALL", 2024, 4800, "official-curated"],
    ]
    for row in gdp_metrics_data:
        ws.append(row)
    apply_formatting(ws, gdp_metrics_data)

    # Sheet 9: IO_Multipliers
    ws = wb.create_sheet("IO_Multipliers")
    io_multipliers_data = [
        ["subsector_id", "multiplier_type", "value", "source_tier"],
        ["ALL", "employment", 1.85, "official-curated"],
        ["ALL", "GDP", 1.72, "official-curated"],
        ["ALL", "output", 2.10, "official-curated"],
    ]
    for row in io_multipliers_data:
        ws.append(row)
    apply_formatting(ws, io_multipliers_data)

    # Sheet 10: Goods_Exports
    ws = wb.create_sheet("Goods_Exports")
    goods_exports_data = [
        ["subsector_id", "destination_country", "year", "value_millions_cad", "source_tier"],
        ["ALL", "All countries", 2019, 650, "official-curated"],
        ["ALL", "All countries", 2020, 620, "official-curated"],
        ["ALL", "All countries", 2021, 710, "official-curated"],
        ["ALL", "All countries", 2022, 780, "official-curated"],
        ["ALL", "All countries", 2023, 840, "official-curated"],
        ["ALL", "All countries", 2024, 890, "official-curated"],
    ]
    for row in goods_exports_data:
        ws.append(row)
    apply_formatting(ws, goods_exports_data)

    # Sheet 11: Research_Activity
    ws = wb.create_sheet("Research_Activity")
    research_activity_data = [
        ["subsector_id", "phase", "therapeutic_area", "year", "trial_count", "source_tier"],
        ["ALL", "ALL", "ALL", 2019, 280, "public-direct"],
        ["ALL", "ALL", "ALL", 2020, 265, "public-direct"],
        ["ALL", "ALL", "ALL", 2021, 310, "public-direct"],
        ["ALL", "ALL", "ALL", 2022, 340, "public-direct"],
        ["ALL", "ALL", "ALL", 2023, 375, "public-direct"],
        ["ALL", "ALL", "ALL", 2024, 402, "public-direct"],
    ]
    for row in research_activity_data:
        ws.append(row)
    apply_formatting(ws, research_activity_data)

    # Sheet 12: Funding_Projects
    ws = wb.create_sheet("Funding_Projects")
    funding_projects_data = [
        ["subsector_id", "program_name", "funder", "year", "amount_millions_cad", "source_tier"],
        ["ALL", "CIHR Operating Grants", "Federal", 2024, 45.2, "public-direct"],
        ["ALL", "NRC IRAP", "Federal", 2024, 28.7, "public-direct"],
        ["ALL", "NSERC Research Grants", "Federal", 2024, 18.4, "public-direct"],
        ["ALL", "BC Life Sciences Fund", "Provincial", 2024, 35.0, "official-curated"],
        ["ALL", "Genome BC Projects", "Provincial", 2024, 22.1, "official-curated"],
    ]
    for row in funding_projects_data:
        ws.append(row)
    apply_formatting(ws, funding_projects_data)

    # Sheet 13: Region_Context
    ws = wb.create_sheet("Region_Context")
    region_context_data = [
        ["region_name", "population", "life_sciences_share_pct", "source_tier"],
        ["Metro Vancouver", 2800000, 0.59, "official-curated"],
        ["Victoria", 420000, 0.67, "official-curated"],
        ["Kelowna", 230000, 0.48, "official-curated"],
        ["Prince George", 98000, 0.61, "official-curated"],
    ]
    for row in region_context_data:
        ws.append(row)
    apply_formatting(ws, region_context_data)

    # Sheet 14: Talent_Pipeline
    ws = wb.create_sheet("Talent_Pipeline")
    talent_pipeline_data = [
        ["subsector_id", "institution", "credential_type", "year", "graduates_count", "source_tier"],
        ["ALL", "ALL", "ALL", 2019, 2850, "official-curated"],
        ["ALL", "ALL", "ALL", 2020, 2920, "official-curated"],
        ["ALL", "ALL", "ALL", 2021, 3100, "official-curated"],
        ["ALL", "ALL", "ALL", 2022, 3280, "official-curated"],
        ["ALL", "ALL", "ALL", 2023, 3450, "official-curated"],
        ["ALL", "ALL", "ALL", 2024, 3620, "official-curated"],
    ]
    for row in talent_pipeline_data:
        ws.append(row)
    apply_formatting(ws, talent_pipeline_data)

    # Sheet 15: Strategy_Goals
    ws = wb.create_sheet("Strategy_Goals")
    strategy_goals_data = [
        ["goal_id", "goal_name", "pillar", "rag_status", "progress_pct", "target_value", "target_unit", "target_year", "baseline_year", "baseline_value", "related_subsectors"],
        ["LW-G01", "Grow sector employment to 30,000 FTE by 2030", "Workforce", "GREEN", 74, 30000, "FTE", 2030, 2019, 18500, "ALL"],
        ["LW-G02", "Double life sciences exports to $1.5B by 2030", "Trade", "AMBER", 59, 1500, "millions_cad", 2030, 2019, 650, "ALL"],
        ["LW-G03", "Increase GDP contribution to $7B by 2030", "Economic Growth", "GREEN", 69, 7000, "millions_cad", 2030, 2019, 3900, "ALL"],
        ["LW-G04", "Grow business count to 1,500 by 2030", "Investment", "AMBER", 64, 1500, "count", 2030, 2019, 980, "ALL"],
        ["LW-G05", "Attract $500M in new investment annually", "Investment", "RED", 42, 500, "millions_cad", 2030, 2019, 210, "ALL"],
        ["LW-G06", "Establish 3 new anchor company HQs in BC", "Investment", "AMBER", 33, 3, "companies", 2030, 2019, 0, "ALL"],
        ["LW-G07", "Grow active clinical trials to 600 by 2030", "Research", "GREEN", 67, 600, "trials", 2030, 2019, 280, "CRO"],
        ["LW-G08", "Increase life sciences graduates by 25%", "Workforce", "GREEN", 82, 3562, "graduates", 2030, 2019, 2850, "ALL"],
    ]
    for row in strategy_goals_data:
        ws.append(row)
    apply_formatting(ws, strategy_goals_data)

    # Sheet 16: Strategy_Initiatives
    ws = wb.create_sheet("Strategy_Initiatives")
    strategy_initiatives_data = [
        ["initiative_id", "initiative_name", "pillar", "rag_status", "status", "progress_pct", "owner_ministry", "next_milestone", "next_milestone_date", "last_review_date", "related_subsectors", "description"],
        ["LW-I01", "BC Life Sciences Fund $100M allocation", "Investment", "GREEN", "Active", 85, "Ministry of Jobs", "Final disbursements", "2025-06-30", "2025-01-15", "ALL", "Deploy $100M provincial fund to attract and retain life sciences investment in BC"],
        ["LW-I02", "Genome BC Strategic Partnership Renewal", "Research", "AMBER", "Active", 60, "Ministry of Health", "Partnership agreement signed", "2025-09-30", "2025-01-10", "BIOTECH", "Renew 5-year strategic partnership with Genome BC for precision medicine initiatives"],
        ["LW-I03", "Life Sciences Talent Attraction Program", "Workforce", "AMBER", "Active", 45, "Ministry of Post-Secondary", "Program launch", "2025-04-30", "2024-12-01", "ALL", "Attract international life sciences talent through targeted immigration and recruitment"],
        ["LW-I04", "BC Clinical Trials Network Expansion", "Research", "GREEN", "Active", 78, "Ministry of Health", "Network expansion complete", "2025-03-31", "2025-01-20", "CRO", "Expand BC Clinical Trials Network to include 5 additional regional sites"],
        ["LW-I05", "Regulatory Fast-Track Pilot Program", "Policy", "RED", "Planning", 25, "Ministry of Health", "Pilot terms finalized", "2025-12-31", "2024-11-15", "ALL", "Pilot streamlined regulatory review process for innovative life sciences products"],
        ["LW-I06", "UBC-Industry Research Partnership Program", "Research", "GREEN", "Active", 90, "Ministry of Advanced Education", "Year 2 funding confirmed", "2025-07-01", "2025-01-18", "BIOTECH|PHARMA", "Multi-year research partnership between UBC and 12 industry partners"],
    ]
    for row in strategy_initiatives_data:
        ws.append(row)
    apply_formatting(ws, strategy_initiatives_data)

    # Sheet 17: Risks_Gaps
    ws = wb.create_sheet("Risks_Gaps")
    risks_gaps_data = [
        ["risk_name", "category", "description", "severity", "likelihood", "rag_status", "owner", "mitigation", "last_reviewed"],
        ["Talent shortage in specialized roles", "Workforce", "Growing global competition for life sciences PhDs and specialized lab technicians may constrain BC sector growth", "HIGH", "Likely", "AMBER", "Ministry of Post-Secondary", "Expand international recruitment; partner with PSIs on co-op programs", "2025-01-15"],
        ["Regulatory uncertainty for novel therapies", "Policy", "Federal regulatory framework for gene therapies and AI-driven diagnostics remains unclear", "HIGH", "Possible", "AMBER", "Ministry of Health", "Engage Health Canada; join national regulatory working group", "2025-01-10"],
        ["Limited wet lab and biomanufacturing space", "Infrastructure", "BC faces chronic shortage of purpose-built wet lab facilities especially in Metro Vancouver", "CRITICAL", "Likely", "RED", "Ministry of Jobs", "Commission lab space study; explore public-private lab hubs", "2024-12-01"],
        ["Concentration risk in Metro Vancouver", "Geographic", "Over 75% of sector employment in one metro area creates resilience risk", "MEDIUM", "Possible", "GREEN", "Ministry of Regional Development", "Invest in regional life sciences clusters in Kelowna and Victoria", "2025-01-20"],
        ["Data gaps in smaller subsectors", "Data & Measurement", "Reliable employment and revenue data unavailable for CRO and Other/Emerging subsectors", "MEDIUM", "Certain", "AMBER", "BC Stats", "Commission targeted sector survey; supplement with survey data", "2024-11-15"],
    ]
    for row in risks_gaps_data:
        ws.append(row)
    apply_formatting(ws, risks_gaps_data)

    # Sheet 18: Evidence
    ws = wb.create_sheet("Evidence")
    evidence_data = [
        ["type", "title", "summary", "author", "publication_date", "confidence_tag", "source_tier", "related_goal_id", "related_initiative_id", "link"],
        ["Research Report", "BC Life Sciences Sector Profile 2024", "Comprehensive profile of BC life sciences sector including employment, revenue, and regional distribution data", "BC Stats", "2024-09-15", "high", "official-curated", "LW-G01|LW-G03", "", "https://www2.gov.bc.ca/gov/content/data/statistics"],
        ["Policy Document", "Look West Strategy 2030 Framework", "Full strategy document outlining 8 goals and 24 initiatives for BC life sciences sector development", "Ministry of Jobs", "2023-11-01", "high", "official-curated", "LW-G01|LW-G02|LW-G03", "LW-I01", "https://www2.gov.bc.ca/gov/content/employment-business/economic-development"],
        ["Data Release", "BC Clinical Trials Registry Q4 2024", "Quarterly update on active clinical trials with a BC site from ClinicalTrials.gov", "Health Canada", "2025-01-10", "medium", "public-direct", "LW-G07", "LW-I04", "https://clinicaltrials.gov"],
        ["Research Report", "Life Sciences Labour Market Outlook 2024-2030", "Forecasts demand for life sciences workers in BC through 2030 across subsectors", "WorkBC", "2024-06-30", "medium", "public-prepared", "LW-G01", "LW-I03", "https://www.workbc.ca"],
    ]
    for row in evidence_data:
        ws.append(row)
    apply_formatting(ws, evidence_data)

    # Sheet 19: Manual_Overrides (headers only)
    ws = wb.create_sheet("Manual_Overrides")
    manual_overrides_data = [
        ["metric_id", "subsector_id", "province", "year", "value", "unit", "override_reason", "approved_by", "approved_date"],
    ]
    for row in manual_overrides_data:
        ws.append(row)
    apply_formatting(ws, manual_overrides_data)

    # Sheet 20: Anchor_Companies (headers only)
    ws = wb.create_sheet("Anchor_Companies")
    anchor_companies_data = [
        ["company_name", "subsector_id", "hq_city", "employees_bc", "founded_year", "notes"],
    ]
    for row in anchor_companies_data:
        ws.append(row)
    apply_formatting(ws, anchor_companies_data)

    # Sheet 21: Whats_New
    ws = wb.create_sheet("Whats_New")
    whats_new_data = [
        ["date", "category", "headline", "blurb", "priority", "link"],
        ["2025-01-15", "Data Update", "Q4 2024 Sector Metrics Updated", "Employment figures updated to reflect Q4 2024 BC Stats release. Sector employment now at 22,100 FTE.", "MEDIUM", ""],
        ["2025-01-10", "Milestone", "Clinical Trials Network Expansion Complete", "BC Clinical Trials Network has reached 402 active trials with a BC site, on track for 2030 goal of 600.", "HIGH", ""],
    ]
    for row in whats_new_data:
        ws.append(row)
    apply_formatting(ws, whats_new_data)

    # Save the workbook
    output_path = "/sessions/jolly-sweet-gauss/mnt/BCLS/bc-economy/data/sectors/life_sciences/Life_Sciences.xlsx"
    wb.save(output_path)

    return output_path

if __name__ == "__main__":
    output_path = create_excel()
    print(f"Excel file created successfully: {output_path}")

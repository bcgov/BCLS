"""Generate the Look West per-metric proposal template (fillable .xlsx)."""
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

NAVY = "FF003366"
GOLD = "FFFCBA19"
LIGHT = "FFF1F5F9"
WHITE = "FFFFFFFF"

thin = Side(style="thin", color="FFCBD5E1")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# (header, column width, optional dropdown list)
COLUMNS = [
    ("Metric ID", 10, None),
    ("Stream", 16, None),
    ("Pillar", 16, None),
    ("Theme", 16, None),
    ("Goal / Target", 30, None),
    ("Proposed Metric", 30, None),
    ("What it measures", 34, None),
    ("Relevance (why it fits this goal)", 34, None),
    ("Reliable data exists?", 16, ["Yes", "Partial", "No", "Unknown"]),
    ("Public source available?", 18,
     ["Yes - StatsCan", "Yes - BC Stats", "Yes - other public", "No", "Unknown"]),
    ("Internal input/validation needed?", 20, ["Yes", "No", "Maybe"]),
    ("Proposed data source", 26, None),
    ("Responsible owner / team", 22, None),
    ("Update frequency", 16,
     ["Monthly", "Quarterly", "Semi-annual", "Annual", "Ad hoc", "TBD"]),
    ("Baseline", 14, None),
    ("Target value", 14, None),
    ("Limitations / gaps / risks", 34, None),
    ("Status", 16,
     ["Draft", "Consulted", "Approved", "Needs data", "Rejected"]),
    ("Notes", 30, None),
]

EXAMPLES = [
    ["M-001", "Trade & Investment", "Market Diversification", "Asia-Pacific Exports",
     "Grow BC exports to Asia-Pacific markets",
     "BC merchandise exports to Asia-Pacific (annual $ and YoY %)",
     "Dollar value and growth rate of BC goods exported to Asia-Pacific economies",
     "Directly measures progress on diversifying trade toward priority markets",
     "Yes", "Yes - StatsCan", "No",
     "Statistics Canada - International Trade (Asia-Pacific filter)",
     "Daniel / Trade analyst", "Quarterly", "TBD", "TBD",
     "Commodity classification & destination coding lags ~2 months",
     "Draft", "Confirm baseline year with team"],
    ["M-002", "Sector Growth", "Clean Economy", "Clean-tech Investment",
     "Attract clean-economy investment to BC",
     "Announced clean-economy investment ($) tied to Look West",
     "Total announced investment value in clean-tech/clean-economy projects",
     "Tracks investment momentum the strategy aims to catalyze",
     "Partial", "No", "Yes",
     "Internal funding/announcement tracker + gov't announcements",
     "Investment promotion team", "Quarterly", "TBD", "TBD",
     "Announcements are not actuals; needs internal validation of realized $",
     "Draft", "Distinguish announced vs. realized investment"],
]


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True, size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border
    ws.row_dimensions[1].height = 38


def main(dst):
    wb = Workbook()

    # ---- Sheet 1: Instructions ----
    info = wb.active
    info.title = "Instructions"
    info.sheet_view.showGridLines = False
    info.column_dimensions["A"].width = 4
    info.column_dimensions["B"].width = 100

    rows = [
        ("title", "Look West Strategy - Metric Proposal Template"),
        ("sub", "Fillable worksheet - draft one row per proposed metric, then consult & seek approval"),
        ("blank", ""),
        ("h", "Purpose"),
        ("p", "For every Look West goal/target, propose one or more practical, reliable metrics. "
              "Complete one row per metric on the 'Metric Proposals' tab. Once drafted, metrics are "
              "consulted with internal teams (CP1) and approved by leadership (CP2) before they are "
              "locked into the Look West Data Hub."),
        ("blank", ""),
        ("h", "How to fill each column"),
        ("li", "Metric ID - simple unique code (M-001, M-002, ...)."),
        ("li", "Stream / Pillar / Theme - position in the Look West hierarchy (use the controlled lists from the Hub)."),
        ("li", "Goal / Target - the specific Look West goal this metric serves."),
        ("li", "Proposed Metric - short name of the metric."),
        ("li", "What it measures - plain-language definition."),
        ("li", "Relevance - why this metric is meaningful for THIS goal."),
        ("li", "Reliable data exists? - Yes / Partial / No / Unknown."),
        ("li", "Public source available? - StatsCan, BC Stats, other public, or No."),
        ("li", "Internal input/validation needed? - whether an internal team must provide or confirm the data."),
        ("li", "Proposed data source - the specific dataset/source you expect to use."),
        ("li", "Responsible owner / team - who supplies or validates the data."),
        ("li", "Update frequency - how often the metric should refresh."),
        ("li", "Baseline / Target value - starting point and target, where known (use TBD if not yet set)."),
        ("li", "Limitations / gaps / risks - known data issues, lags, or caveats."),
        ("li", "Status - Draft -> Consulted -> Approved (or Needs data / Rejected)."),
        ("li", "Notes - anything else reviewers should know."),
        ("blank", ""),
        ("h", "Workflow"),
        ("p", "Draft (Mehdi + Daniel)  ->  Consult internal teams/SMEs (CP1)  ->  Revise  ->  "
              "Leadership approval (CP2, Jacqueline)  ->  Populate the Data Hub."),
        ("blank", ""),
        ("note", "Two example rows are pre-filled on the 'Metric Proposals' tab - overwrite or delete them."),
    ]

    r = 1
    for kind, text in rows:
        cell = info.cell(row=r, column=2, value=text)
        if kind == "title":
            cell.font = Font(bold=True, size=16, color=NAVY)
            info.row_dimensions[r].height = 24
        elif kind == "sub":
            cell.font = Font(italic=True, size=10, color="FF64748B")
        elif kind == "h":
            cell.font = Font(bold=True, size=12, color=NAVY)
        elif kind == "li":
            cell.value = "  -  " + text
            cell.font = Font(size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        elif kind == "note":
            cell.font = Font(italic=True, size=10, color="FFB45309")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        else:  # p
            cell.font = Font(size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            info.row_dimensions[r].height = 46
        r += 1

    # ---- Sheet 2: Metric Proposals ----
    ws = wb.create_sheet("Metric Proposals")
    ws.sheet_view.showGridLines = False
    headers = [c[0] for c in COLUMNS]
    ws.append(headers)
    for idx, (name, width, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    style_header(ws, len(COLUMNS))

    for ex in EXAMPLES:
        ws.append(ex)

    # body styling for a generous number of blank rows
    last_row = 200
    for row in range(2, last_row + 1):
        ws.row_dimensions[row].height = 30
        fill = LIGHT if row % 2 == 0 else WHITE
        for col in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(size=10)

    # dropdowns
    for idx, (_, _, choices) in enumerate(COLUMNS, start=1):
        if not choices:
            continue
        letter = ws.cell(row=1, column=idx).column_letter
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(choices) + '"',
            allow_blank=True,
        )
        dv.add(f"{letter}2:{letter}{last_row}")
        ws.add_data_validation(dv)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(COLUMNS)).column_letter}1"

    wb.save(dst)
    print(f"Saved {dst}")


if __name__ == "__main__":
    main(sys.argv[1])

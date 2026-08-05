# -*- coding: utf-8 -*-
"""
Build a feature-inventory workbook for the BCLS dashboard hub.
One sheet per dashboard page. Each sheet lists every feature
(KPI / chart / table / map / control), the data it shows,
the data source, the underlying data file, and availability.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\mehdi\bcls\BCLS_Dashboard_Feature_Inventory.xlsx"

# ---- styling ---------------------------------------------------------------
NAVY = "003366"
GOLD = "FCBA19"
HDR_FILL = PatternFill("solid", fgColor=NAVY)
SUB_FILL = PatternFill("solid", fgColor="EEF2F7")
LIVE_FILL = PatternFill("solid", fgColor="E4F4E8")
PLAN_FILL = PatternFill("solid", fgColor="FBEFE0")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color=NAVY, size=15)
SUB_FONT = Font(italic=True, color="555555", size=10)
CELL_FONT = Font(size=10)
thin = Side(style="thin", color="D0D7E2")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = ["#", "Feature", "Type", "What it shows / Data",
        "Data Source", "Data File / Reference", "Availability", "Notes"]
WIDTHS = [4, 34, 14, 46, 30, 34, 16, 40]

def style_sheet(ws, title, subtitle, rows):
    ws.sheet_view.showGridLines = False
    # title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    c = ws.cell(1, 1, title); c.font = TITLE_FONT; c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLS))
    c = ws.cell(2, 1, subtitle); c.font = SUB_FONT
    ws.row_dimensions[2].height = 16
    # header row
    hr = 4
    for j, name in enumerate(COLS, 1):
        cc = ws.cell(hr, j, name); cc.fill = HDR_FILL; cc.font = HDR_FONT
        cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cc.border = BORDER
    ws.row_dimensions[hr].height = 26
    # data rows
    for i, row in enumerate(rows, 1):
        r = hr + i
        avail = row[5] if len(row) > 5 else ""
        fill = LIVE_FILL if str(avail).lower().startswith("live") else (
               PLAN_FILL if any(k in str(avail).lower() for k in ("coming", "planned", "inactive", "not")) else None)
        vals = [i] + list(row)
        for j, v in enumerate(vals, 1):
            cc = ws.cell(r, j, v); cc.font = CELL_FONT; cc.border = BORDER
            cc.alignment = Alignment(vertical="top", wrap_text=True,
                                     horizontal="center" if j in (1, 7) else "left")
            if j == 7 and fill:
                cc.fill = fill
    for j, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A5"

# Each data row: [Feature, Type, What/Data, Source, File/Ref, Availability, Notes]

wb = openpyxl.Workbook()

# ============================ OVERVIEW ======================================
ws = wb.active; ws.title = "Overview"
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:F1"); c = ws.cell(1,1,"BCLS Dashboard - Feature & Data Inventory"); c.font = TITLE_FONT
ws.merge_cells("A2:F2"); c = ws.cell(2,1,"BC / Look West single hub dashboard. One sheet per page. Generated from the live HTML in dashboard/.")
c.font = SUB_FONT
ov_cols = ["#", "Dashboard Page", "Hub Nav / Route", "Source File", "Availability", "Summary"]
ov_w = [4, 30, 26, 46, 16, 52]
hr = 4
for j,name in enumerate(ov_cols,1):
    cc=ws.cell(hr,j,name); cc.fill=HDR_FILL; cc.font=HDR_FONT
    cc.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cc.border=BORDER
ws.row_dimensions[hr].height=26
overview = [
 ["Hub (shell)", "Landing / navigation", "dashboard/hub/html/dashboard.html", "Live",
  "Parent shell. Top nav + sector card grid; embeds all child pages via iframe."],
 ["BC Economy Snapshot", "?section=macro", "dashboard/snapshot/html/dashboard.html", "Live",
  "GDP, labour market, inflation and trade for BC. Live Statistics Canada data."],
 ["Look West Strategy Monitor", "?section=strategy", "dashboard/strategy-monitor/html/dashboard.html", "Live",
  "Strategy delivery KPIs, media coverage, investment pipeline, policy, data inventory."],
 ["Look West Tracker (hub)", "?section=tracker", "dashboard/tracker/html/dashboard.html", "Live",
  "Card hub linking to 6 tracker pages (2 active, 4 inactive)."],
 ["  - Targets Tracker", "tracker page: targets", "tracker/html/pages/lw_targets_tracker.html", "Live (Active)",
  "Outcome targets, baselines, progress; 14 charts + tables."],
 ["  - Goals Tracker", "tracker page: goals", "tracker/html/pages/lw_goals_tracker.html", "Live (Active)",
  "Goal-level metrics and progress narratives grouped by target."],
 ["  - Policy & Regulations Tracker", "tracker page: policy-regulation", "tracker/html/pages/lw_policy_regulations_tracker.html", "Built (Inactive)",
  "Policy/regulation lifecycle. Built but disabled in hub."],
 ["  - Funding & Investment Tracker", "tracker page: funding-investment", "tracker/html/pages/lw_funding_investment_tracker.html", "Built (Inactive)",
  "Funding programs + transactions. Built but disabled in hub."],
 ["  - Infrastructure Tracker", "tracker page: infrastructure", "tracker/html/pages/lw_infrastructure_tracker.html", "Built (Inactive)",
  "Infrastructure delivery pipeline. Built but disabled in hub."],
 ["  - Investment Promotion Tracker", "tracker page: investment-promotion", "tracker/html/pages/lw_investment_promotion_tracker.html", "Built (Inactive)",
  "Investment promotion pipeline. Built but disabled in hub."],
 ["Projects", "?section=projects", "dashboard/projects/html/dashboard.html", "Live",
  "BC major projects map, investment/GDP chart, projects + tracker tables."],
 ["Sectors (hub grid)", "?section=sectors", "dashboard/hub (sector card grid)", "Live",
  "10 sector cards; only Life Sciences and Tourism are enabled."],
 ["  - Life Sciences", "sector-detail: life-sciences", "sectors/life_sciences/html/dashboard.html", "Live",
  "Full sector dashboard: KPIs, charts, strategy goals, commitments, evidence."],
 ["  - Tourism", "sector-detail: tourism", "sectors/tourism/html/dashboard.html", "Live",
  "Sector dashboard: subsector GDP/business/labour charts, goals, announcements."],
 ["  - 8 other sectors", "sector-detail (disabled)", "sectors/<name>/html/dashboard.html", "Coming Soon",
  "Trade & Logistics, Maritime, AI & Quantum, Aerospace, Construction Innovation, Agriculture, Future 01/02 - placeholder shells."],
]
for i,row in enumerate(overview,1):
    r=hr+i; avail=row[3]
    fill = LIVE_FILL if avail.lower().startswith("live") else PLAN_FILL if any(k in avail.lower() for k in ("coming","built","inactive")) else None
    for j,v in enumerate([i]+row,1):
        cc=ws.cell(r,j,v); cc.font=CELL_FONT; cc.border=BORDER
        cc.alignment=Alignment(vertical="top",wrap_text=True,horizontal="center" if j in(1,) else "left")
        if j==5 and fill: cc.fill=fill
for j,w in enumerate(ov_w,1):
    ws.column_dimensions[get_column_letter(j)].width=w
ws.freeze_panes="A5"

# ============================ HUB ===========================================
rows = [
 ["Top navigation bar", "Nav control", "Buttons: BC Economy Snapshot, Look West Strategy Monitor, Look West Goals Tracker, More (dropdown)", "Inline JS (SECTIONS)", "hub/dashboard.html", "Live", "Routes via ?section= query param."],
 ["\"More\" dropdown", "Nav control", "Jumps to Sectors, Projects and other sections", "Inline JS", "hub/dashboard.html", "Live", ""],
 ["BC Economy Snapshot panel", "Embedded page (iframe)", "Embeds the BC Economy Snapshot dashboard", "Child page", "snapshot/html/dashboard.html", "Live", "Auto-fits iframe height."],
 ["Look West Strategy panel", "Embedded page (iframe)", "Embeds the Strategy Monitor dashboard", "Child page", "strategy-monitor/html/dashboard.html", "Live", ""],
 ["Look West Tracker panel", "Embedded page (iframe)", "Embeds the Tracker hub", "Child page", "tracker/html/dashboard.html", "Live", ""],
 ["Sectors hub (card grid)", "Card grid", "One card per BC industry sector with status + last-updated", "Inline JS (SECTORS registry)", "hub/dashboard.html", "Live", "8 visible cards; Life Sciences + Tourism enabled, rest planned."],
 ["Sector detail panel", "Embedded page (iframe)", "Embeds the selected sector's dashboard", "Child page", "sectors/<sector>/html/dashboard.html", "Live", "Breadcrumb back to All Sectors."],
 ["Projects panel", "Embedded page (iframe)", "Embeds the Projects dashboard", "Child page", "projects/html/dashboard.html", "Live", ""],
]
style_sheet(wb.create_sheet("Hub"), "Hub (Parent Shell)",
            "dashboard/hub/html/dashboard.html - navigation shell that embeds all child pages via iframe", rows)

# ==================== BC ECONOMY SNAPSHOT ===================================
SC = "Statistics Canada Web Data Service (live API via local proxy scripts/serve.py)"
rows = [
 ["GDP (Current $)", "KPI card", "Latest BC GDP, real & nominal change", SC, "StatCan table 36-10-0711-01", "Live", ""],
 ["Inflation (CPI)", "KPI card", "BC CPI with MoM / YoY change", SC, "StatCan table 18-10-0004-01", "Live", ""],
 ["Employment", "KPI card", "BC employment level, MoM / YoY", SC, "StatCan table 14-10-0355-01", "Live", ""],
 ["Unemployment Rate", "KPI card", "BC unemployment rate, MoM / YoY", SC, "StatCan table 14-10-0355-01", "Live", ""],
 ["Labour Force", "KPI card", "BC labour force, MoM / YoY", SC, "StatCan table 14-10-0355-01", "Live", ""],
 ["International Exports", "KPI card", "BC international goods exports, YoY", SC, "StatCan table 36-10-0709-01", "Live", ""],
 ["International Imports", "KPI card", "BC international goods imports, YoY", SC, "StatCan table 36-10-0709-01", "Live", ""],
 ["Trade Balance", "KPI card", "BC trade balance, YoY", SC, "StatCan table 36-10-0709-01", "Live", ""],
 ["Employment - British Columbia", "Line/bar chart", "Employment trend for BC", SC, "StatCan table 14-10-0355-01", "Live", "canvas chart-labour"],
 ["GDP - British Columbia", "Line chart", "GDP trend for BC", SC, "StatCan table 36-10-0711-01", "Live", "canvas chart-gdp"],
 ["GDP Share by Industry - BC", "Bar/pie chart", "Industry composition of BC GDP", SC, "StatCan table 36-10-0402-01 / 36-10-0711", "Live", "canvas chart-gdp-share"],
 ["Labour market detail table", "Table", "Month, goods/services, industry, employment, change, growth, rank", SC, "StatCan table 14-10-0355-01", "Live", ""],
 ["Consumer Price Index - BC", "Line chart", "CPI trend for BC", SC, "StatCan table 18-10-0004-01", "Live", "canvas chart-cpi"],
 ["Contributions to YoY CPI Inflation", "Stacked bar chart", "Component contributions to BC inflation", SC, "StatCan table 18-10-0004-01", "Live", "canvas chart-cpi-contrib"],
 ["Interprovincial Trade - BC", "Chart", "Interprovincial trade flows for BC", SC, "StatCan (interprovincial trade table)", "Live", "canvas chart-interprov-trade"],
 ["International Trade - BC", "Chart", "International trade flows for BC", SC, "StatCan table 36-10-0709-01", "Live", "canvas chart-international-trade"],
 ["Update Data button", "Control", "Refreshes live StatCan pulls through the local proxy", "Local proxy", "serve.py /api/statcan-wds, /api/statcan-csv", "Live", ""],
]
style_sheet(wb.create_sheet("BC Economy Snapshot"), "BC Economy Snapshot",
            "dashboard/snapshot/html/dashboard.html - macro panel. Live Statistics Canada data via local proxy. (File also retains legacy strategy/sectors/projects panels.)", rows)

# ==================== STRATEGY MONITOR ======================================
rows = [
 ["Strategy Delivery Snapshot - hero KPIs", "KPI cards", "Jobs created & protected, total funding programs value, and other strategic outcome tiles", "Look West CSV datasets", "Multiple LW CSVs (see below)", "Live", "kpi-hero grid"],
 ["Data coverage - mini KPIs", "KPI cards", "Counts: Action Plans, Targets, Goals, Metrics Monitored, Announcements, Policy & Regulation, Funding Programs, Infrastructures, Investment Promotions", "Look West CSV datasets", "Action Plans.csv, Targets.csv, LW Goals.csv, etc.", "Live", "kpi-mini grid"],
 ["Investment & Project Pipeline charts", "Chart gallery (toggleable)", "Funding by source ($M) [bar+donut], commitment by program, infrastructures value by category/region, investment promotion by sector/status, major projects by status, commitment & jobs by project, policy by type, targets by pillar/sector", "Look West CSV datasets", "Funding Programs.csv, Infrastructures.csv, Investment Promotion.csv, Major Projects.csv, Policy & Regulations.csv, Targets.csv", "Live", "14 charts, user show/hide"],
 ["Media Coverage - Announcements table", "Filterable table", "Date, Source, Pillar, Theme, Headline, Primary Sector, Region", "Look West CSV", "LW Announcements.csv", "Live", "Filters: source/pillar/theme + search; sortable; paged"],
 ["Media Coverage - News table", "Filterable table", "Date, Source, Pillar, Theme, Title, Sector, Region", "Look West CSV", "LW News.csv", "Live", "Filters + search; sortable; paged"],
 ["Investment & Project Pipeline table", "Filterable table", "Date, Project/Program, Pillar, Theme, Provincial/Attracted/Federal/Other/Total ($M), Jobs Created, Region", "Look West CSV", "Funding Programs.csv, Major Projects.csv, Infrastructures.csv, Investment Promotion.csv", "Live", "Filters: pillar/theme/funding source + search"],
 ["Policy & Regulations table", "Filterable table", "Date Announced, Policy/Regulation, Jurisdiction, LW Sector, Policy Type, LW Outcome Area, Source", "Look West CSV", "Policy & Regulations.csv", "Live", "Filters: jurisdiction/sector/type + search"],
 ["Data Inventory", "Expandable lists", "Every captured record grouped by dataset (Action Plans, Targets, Goals, Metrics, Announcements, Policy, Funding, Infrastructures, Investment Promotions)", "Look West CSV datasets", "All LW CSVs", "Live", "Accordion by dataset"],
 ["Update Data button", "Control", "Reloads the CSV datasets", "Local CSV files", "data/*.csv", "Live", ""],
]
style_sheet(wb.create_sheet("Strategy Monitor"), "Look West Strategy Monitor",
            "dashboard/strategy-monitor/html/dashboard.html - strategy delivery KPIs, media, pipeline, policy and data inventory. Source: Look West CSV datasets in data/.", rows)

# ==================== TRACKER HUB ==========================================
rows = [
 ["Tracker card grid", "Card hub", "One card per tracker stream with title, subtitle, description and Active/Inactive state", "Inline JS (TRACKER_PAGES)", "tracker/dashboard.html", "Live", "6 cards"],
 ["Targets Tracker card", "Link card", "Opens LW Targets Tracker", "-", "pages/lw_targets_tracker.html", "Live (Active)", ""],
 ["Goals Tracker card", "Link card", "Opens LW Goals Tracker", "-", "pages/lw_goals_tracker.html", "Live (Active)", ""],
 ["Policy & Regulations card", "Link card", "Policy/regulation lifecycle", "-", "pages/lw_policy_regulations_tracker.html", "Inactive", "Disabled in hub"],
 ["Funding & Investment card", "Link card", "Program allocations & transactions", "-", "pages/lw_funding_investment_tracker.html", "Inactive", "Disabled in hub"],
 ["Infrastructure card", "Link card", "Delivery pipeline & milestones", "-", "pages/lw_infrastructure_tracker.html", "Inactive", "Disabled in hub"],
 ["Investment Promotion card", "Link card", "Pipeline conversion & outcomes", "-", "pages/lw_investment_promotion_tracker.html", "Inactive", "Disabled in hub"],
 ["Detail frame", "Embedded page (iframe)", "Loads the selected tracker page", "Child page", "pages/lw_*_tracker.html", "Live", "Routes via ?page="],
]
style_sheet(wb.create_sheet("Tracker Hub"), "Look West Tracker (Hub)",
            "dashboard/tracker/html/dashboard.html - card hub linking to 6 tracker pages (2 active, 4 inactive).", rows)

# ==================== TARGETS TRACKER ======================================
TT = "Look West CSV: Targets.csv + Target Tracking.csv"
rows = [
 ["Summary KPI cards", "KPI cards", "Target counts and tracked-activity coverage", TT, "Targets.csv, Target Tracking.csv", "Live", ""],
 ["Targets with Limited Tracked Activity", "Insight list", "Targets with few/no linked tracker rows", TT, "Target Tracking.csv", "Live", "insight-card"],
 ["Tracker Updates by Sector", "Bar chart", "Mapped updates grouped by sector", TT, "Target Tracking.csv", "Live", "chart-sector-updates"],
 ["Tracker Updates by Stream", "Bar chart", "Updates grouped by Look West stream", TT, "Target Tracking.csv", "Live", "chart-stream-updates"],
 ["Tracker Updates by Pillar", "Bar chart", "Updates grouped by pillar", TT, "Target Tracking.csv", "Live", "chart-pillar-updates"],
 ["Tracker Updates by Theme", "Bar chart", "Updates grouped by theme", TT, "Target Tracking.csv", "Live", "chart-theme-updates"],
 ["Achievement Status Mix", "Doughnut chart", "Status categories from tracker records", TT, "Target Tracking.csv", "Live", "chart-achievement-mix"],
 ["Targets by Stream", "Bar chart", "Targets dataset only", TT, "Targets.csv", "Live", "chart-targets-by-stream"],
 ["Targets by Pillar", "Bar chart", "Targets dataset only", TT, "Targets.csv", "Live", "chart-targets-by-pillar-mix"],
 ["Targets by Section", "Bar chart", "Based on Page field", TT, "Targets.csv", "Live", "chart-targets-by-section"],
 ["Targets by Theme", "Bar chart", "Targets dataset only", TT, "Targets.csv", "Live", "chart-targets-by-theme"],
 ["Targets by Sector", "Bar chart", "Targets dataset only", TT, "Targets.csv", "Live", "chart-targets-by-sector"],
 ["Targets by Region", "Bar chart", "Region fields (if available)", TT, "Targets.csv", "Live", "chart-targets-by-region"],
 ["Targets by Lead Ministry", "Bar chart", "Targets dataset only", TT, "Targets.csv", "Live", "chart-targets-by-lead-ministry"],
 ["Targets by Timeframe", "Bar chart", "Targets dataset only", TT, "Targets.csv", "Live", "chart-targets-by-timeframe"],
 ["Directed vs Collective Targets by Pillar", "Stacked bar chart", "Directed/collective mix by pillar", TT, "Targets.csv", "Live", "chart-targets-directed-by-pillar"],
 ["Targets table", "Table", "Target, Stream, Pillar, Theme, Sector, Timeframe", TT, "Targets.csv", "Live", ""],
 ["Achievement / outcomes table", "Table", "Target, Stream, Achievement, Description, Jobs Created, Jobs Saved, Investment Attracted ($M), Revenue to BC GDP ($M), Revenue to BC Gov ($M)", TT, "Target Tracking.csv", "Live", ""],
 ["Select Charts panel", "Control", "Show/hide charts", "-", "-", "Live", ""],
]
style_sheet(wb.create_sheet("Tracker - Targets"), "LW Targets Tracker",
            "tracker/html/pages/lw_targets_tracker.html - outcome targets, baselines, progress cadence.", rows)

# ==================== GOALS TRACKER ========================================
rows = [
 ["Goal groups (by target)", "Cards / sections", "Look West goals grouped by their linked target", "Look West CSV", "LW Goals.csv + Goals Tracker.csv", "Live", ""],
 ["Per-goal metric summary", "Metric tiles", "Up to 3 metrics per goal (primary highlighted)", "Look West CSV", "Goals Tracker.csv", "Live", ""],
 ["Progress narrative", "Text block", "Progress narrative per goal", "Look West CSV", "LW Goals.csv", "Live", ""],
 ["Summary KPI cards", "KPI cards", "Goal counts / progress rollups", "Look West CSV", "LW Goals.csv, Goals Tracker.csv", "Live", ""],
]
style_sheet(wb.create_sheet("Tracker - Goals"), "LW Goals Tracker",
            "tracker/html/pages/lw_goals_tracker.html - goal-level metrics and progress narratives grouped by target.", rows)

# ==================== POLICY & REG TRACKER (inactive) ======================
PR = "Look West CSV: Policy & Regulations.csv + Policy & Regulations Tracker.csv"
rows = [
 ["Tracker Updates by Stream", "Bar chart", "Policy updates by stream", PR, "Policy & Regulations Tracker.csv", "Inactive", ""],
 ["Tracker Updates by Policy Type", "Bar chart", "Updates by policy type", PR, "Policy & Regulations Tracker.csv", "Inactive", ""],
 ["Tracker Updates by Pillar", "Bar chart", "Updates by pillar", PR, "Policy & Regulations Tracker.csv", "Inactive", ""],
 ["Tracker Updates by Theme", "Bar chart", "Updates by theme", PR, "Policy & Regulations Tracker.csv", "Inactive", ""],
 ["Achievement Status Mix", "Doughnut chart", "Status categories", PR, "Policy & Regulations Tracker.csv", "Inactive", ""],
 ["Policy & Regulations Tracker table", "Table", "Policy/regulation records with status changes, ministries, timing, linked announcements", PR, "Policy & Regulations.csv, Policy & Regulations Tracker.csv", "Inactive", ""],
]
style_sheet(wb.create_sheet("Tracker - Policy & Reg"), "LW Policy & Regulations Tracker (Inactive)",
            "tracker/html/pages/lw_policy_regulations_tracker.html - built but disabled in the hub.", rows)

# ==================== FUNDING & INVESTMENT TRACKER (inactive) ==============
FI = "Look West CSV: Funding Programs.csv + Funding Program Tracker.csv"
rows = [
 ["Transactions by Stream", "Bar chart", "Funding transactions by stream", FI, "Funding Program Tracker.csv", "Inactive", ""],
 ["Transaction Amount by Funding Source", "Bar chart", "Amount by funding source", FI, "Funding Program Tracker.csv", "Inactive", ""],
 ["Transactions by Pillar", "Bar chart", "Transactions by pillar", FI, "Funding Program Tracker.csv", "Inactive", ""],
 ["Transactions by Theme", "Bar chart", "Transactions by theme", FI, "Funding Program Tracker.csv", "Inactive", ""],
 ["Transaction Type Mix", "Doughnut chart", "Transaction type distribution", FI, "Funding Program Tracker.csv", "Inactive", ""],
 ["Funding Program Tracker table", "Table", "Programs from planned envelopes to transaction-level updates: recipients, sectors, regions, outcomes", FI, "Funding Programs.csv, Funding Program Tracker.csv", "Inactive", ""],
]
style_sheet(wb.create_sheet("Tracker - Funding"), "LW Funding & Investment Tracker (Inactive)",
            "tracker/html/pages/lw_funding_investment_tracker.html - built but disabled in the hub.", rows)

# ==================== INFRASTRUCTURE TRACKER (inactive) ====================
IN = "Look West CSV: Infrastructures.csv + Infrastructure Tracker.csv / Infrastructures Transactions.csv"
rows = [
 ["Transactions by Stream", "Bar chart", "Infrastructure transactions by stream", IN, "Infrastructure Tracker.csv", "Inactive", ""],
 ["Transaction Amount by Funding Source", "Bar chart", "Amount by funding source", IN, "Infrastructure Tracker.csv", "Inactive", ""],
 ["Transactions by Pillar", "Bar chart", "Transactions by pillar", IN, "Infrastructure Tracker.csv", "Inactive", ""],
 ["Transactions by Theme", "Bar chart", "Transactions by theme", IN, "Infrastructure Tracker.csv", "Inactive", ""],
 ["Achievement Status Mix", "Doughnut chart", "Status categories", IN, "Infrastructure Tracker.csv", "Inactive", ""],
 ["Infrastructure Transactions table", "Table", "Initiatives across planning/procurement/delivery: milestone movement, transaction history, constraints", IN, "Infrastructures.csv, Infrastructure Tracker.csv", "Inactive", ""],
]
style_sheet(wb.create_sheet("Tracker - Infrastructure"), "LW Infrastructure Tracker (Inactive)",
            "tracker/html/pages/lw_infrastructure_tracker.html - built but disabled in the hub.", rows)

# ==================== INVESTMENT PROMOTION TRACKER (inactive) ==============
IP = "Look West CSV: Investment Promotion.csv + Investment Promotion Tracker.csv"
rows = [
 ["Updates by Stream", "Bar chart", "Investment promotion updates by stream", IP, "Investment Promotion Tracker.csv", "Inactive", ""],
 ["Investment Attracted by Sector", "Bar chart", "Investment attracted by sector", IP, "Investment Promotion Tracker.csv", "Inactive", ""],
 ["Updates by Pillar", "Bar chart", "Updates by pillar", IP, "Investment Promotion Tracker.csv", "Inactive", ""],
 ["Updates by Theme", "Bar chart", "Updates by theme", IP, "Investment Promotion Tracker.csv", "Inactive", ""],
 ["Achievement Status Mix", "Doughnut chart", "Status categories", IP, "Investment Promotion Tracker.csv", "Inactive", ""],
 ["Investment Promotion Tracker table", "Table", "Activity from lead generation to conversion: pipeline changes, ministry actions, realized investment", IP, "Investment Promotion.csv, Investment Promotion Tracker.csv", "Inactive", ""],
]
style_sheet(wb.create_sheet("Tracker - Invest Promo"), "LW Investment Promotion Tracker (Inactive)",
            "tracker/html/pages/lw_investment_promotion_tracker.html - built but disabled in the hub.", rows)

# ==================== PROJECTS =============================================
PJ = "Local CSV: Major Projects.csv + Major Project Tracker.csv"
rows = [
 ["Project Snapshot KPIs", "KPI cards", "Rollup counts / totals across major projects", PJ, "Major Projects.csv", "Live", ""],
 ["Filters", "Controls", "Tracker search + filter controls", "-", "-", "Live", ""],
 ["Project Footprint in B.C.", "Map", "Geographic map of BC major projects", PJ + " + BC boundary geojson", "Major Projects.csv, bc_pr_boundary_2021.geojson", "Live", ""],
 ["Investment and GDP by Project", "Bar chart", "Investment and GDP contribution per project", PJ, "Major Projects.csv", "Live", "canvas project-bars"],
 ["Major Projects table", "Table", "Project ID, Name, Region, Status, Lead Ministry, Date Announced, Details", PJ, "Major Projects.csv", "Live", ""],
 ["Major Project Tracker table", "Table", "Project ID, Ministry, Announcement Type, What Changed, Next Step, Note", PJ, "Major Project Tracker.csv", "Live", ""],
]
style_sheet(wb.create_sheet("Projects"), "Projects",
            "dashboard/projects/html/dashboard.html - BC major projects map, investment/GDP chart and tables.", rows)

# ==================== SECTORS HUB ==========================================
rows = [
 ["Life Sciences", "Sector card -> dashboard", "Pharma, medical devices, biotech, clinical services", "Sector Excel workbook", "Life_Sciences_light.xlsx", "Live", "Enabled; last updated Apr 7 2026"],
 ["Tourism", "Sector card -> dashboard", "Visitor economy, hospitality, destination development", "Sector Excel + LW CSVs", "Tourism.xlsx / sector_tourism.csv", "Live", "Enabled"],
 ["Trade & Logistics", "Sector card", "Ports, logistics, warehousing, supply chain", "Sector Excel (planned)", "Trade_and_Logistics.xlsx", "Coming Soon", "Placeholder shell"],
 ["Maritime", "Sector card", "Marine transport, ocean services, coastal industries", "Sector Excel (planned)", "Maritime.xlsx", "Coming Soon", "Placeholder shell"],
 ["AI & Quantum Computing", "Sector card", "AI, quantum research, advanced computing", "Sector Excel (planned)", "AI_Quantum_Computing.xlsx", "Coming Soon", "Placeholder shell"],
 ["Aerospace", "Sector card", "Aerospace manufacturing, MRO, aviation innovation", "Sector Excel (planned)", "Aerospace.xlsx", "Coming Soon", "Placeholder shell"],
 ["Construction Innovation", "Sector card", "Construction productivity, methods, materials", "Sector Excel (planned)", "Construction_Innovation.xlsx", "Coming Soon", "Placeholder shell"],
 ["Agriculture", "Sector card", "Primary agriculture, agri-tech, value-added", "Sector Excel (planned)", "Agriculture.xlsx", "Coming Soon", "Placeholder shell"],
 ["Future Sector 01 / 02", "Sector card (hidden)", "Reserved placeholders", "-", "Future_Sector_0x.xlsx", "Coming Soon", "Hidden in hub grid"],
]
style_sheet(wb.create_sheet("Sectors Hub"), "Sectors (Hub Grid)",
            "dashboard/hub sector card grid - 10 sectors defined; only Life Sciences and Tourism are enabled.", rows)

# ==================== LIFE SCIENCES =========================================
LS = "Sector Excel workbook (single source)"
LSf = "sectors/life_sciences/data/Life_Sciences_light.xlsx"
rows = [
 ["Sector definition / header", "Header block", "Sector overview and subsector list", LS, LSf + " (SECTOR_STRUCTURE, Sector_Definition)", "Live", ""],
 ["What's New panel", "Panel", "Recent updates for the sector", LS, LSf + " (Whats_New)", "Live", "Shown when data present"],
 ["Sector Snapshot - KPI cards", "KPI cards", "Key sector metrics (configurable, show/hide)", LS, LSf + " (KEY_KPIS)", "Live", "Select KPIs control"],
 ["Sector Snapshot - charts", "Chart gallery", "Sector trend charts driven by chart catalog (GDP, business counts, exports, research/clinical trials, talent pipeline, etc.)", LS, LSf + " (CHART_DATA, CHART_CATALOG)", "Live", "Select Charts control; canvas chart-<id>"],
 ["Look West Strategy - goal summary", "Summary row", "Strategy goal rollup for the sector", LS + " + LW goals", LSf + " (STRATEGY_GOALS) / sector_life_sciences.csv", "Live", ""],
 ["Look West Strategy - goals grid", "Goal cards", "Per-goal strategy detail (show/hide goals)", LS, LSf + " (STRATEGY_GOALS)", "Live", "Select Goals control"],
 ["Commitments table", "Table", "Goal, Stream, Pillar, Theme, Timeframe", "Look West CSV", "Action Plans Commitments.csv", "Live", ""],
 ["Relevant Announcements table", "Table", "Date, Headline/subtitle, Source, Region, Funding, Policy/Regulation, Investment Promotion, Infrastructure", "Look West CSV", "LW Announcements.csv", "Live", ""],
 ["Initiative Tracker", "Table", "Sector initiatives tracking", LS, LSf + " (INITIATIVE_TRACKER)", "Live", ""],
 ["Evidence Library", "Table + export", "Supporting evidence records; CSV export", LS, LSf + " (EVIDENCE_LIBRARY)", "Live", "Export to evidence_export.csv"],
 ["Risks & Gaps", "Table", "Sector risks and gaps", LS, LSf + " (RISKS_GAPS)", "Live", ""],
 ["Load Excel File / Update Data", "Control", "Load the sector workbook (strict single-source)", LS, LSf, "Live", "Accepts .xlsx"],
]
style_sheet(wb.create_sheet("Sector - Life Sciences"), "Sector: Life Sciences",
            "sectors/life_sciences/html/dashboard.html - full sector dashboard. Strict single Excel source.", rows)

# ==================== TOURISM ==============================================
TXL = "Sector Excel workbook: Tourism.xlsx"
rows = [
 ["Subsector definition table", "Table", "Subsector, NAICS Codes, Description", TXL, "Tourism.xlsx", "Live", ""],
 ["Real GDP Trend by Subsector", "Line chart", "Real GDP by tourism subsector over time", TXL, "Tourism.xlsx", "Live", "chart-gdp-real"],
 ["Nominal GDP Trend by Subsector", "Line chart", "Nominal GDP by subsector over time", TXL, "Tourism.xlsx", "Live", "chart-gdp-nominal"],
 ["Business Count by Subsector", "Bar chart", "Business counts by subsector", TXL, "Tourism.xlsx", "Live", "chart-biz-subsector"],
 ["Business Size Mix by Subsector", "Stacked bar chart", "Business size distribution by subsector", TXL, "Tourism.xlsx", "Live", "chart-biz-size"],
 ["Employment Trend in Tourism-related Industries", "Line chart", "Employment trend across tourism industries", TXL, "Tourism.xlsx", "Live", "chart-labour"],
 ["Look West goals (per subsector)", "Goal cards", "Strategy goals mapped to subsectors", "Look West CSV", "sector_tourism.csv", "Live", ""],
 ["Commitments table", "Table", "Goal, Stream, Pillar, Theme, Timeframe", "Look West CSV", "Action Plans Commitments.csv", "Live", ""],
 ["Relevant Announcements table", "Table", "Date, Headline/subtitle, Source, Region, Funding, Policy/Regulation, Investment Promotion, Infrastructure", "Look West CSV", "LW Announcements.csv", "Live", ""],
]
style_sheet(wb.create_sheet("Sector - Tourism"), "Sector: Tourism",
            "sectors/tourism/html/dashboard.html - subsector GDP/business/labour charts plus Look West goals & announcements.", rows)

wb.save(OUT)
print("Saved:", OUT)
print("Sheets:", wb.sheetnames)

"""
Quick local server for the restructured BCLS dashboard workspace.
Run from the BCLS root folder:

    python scripts/serve.py

Then open:  http://localhost:8080/dashboards/bc_dashboard_hub/html/dashboard.html
"""
import http.server
import base64
import hashlib
import json
import os
import socket
import socketserver
import threading
import urllib.error
import urllib.parse
import urllib.request
import webbrowser
import ssl
import subprocess
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

try:
    import msal
except Exception:
    msal = None

PORT = 8080
PREFERRED_PORTS = [8080]
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_HOME_MAP_PATH = os.path.join(os.path.expanduser("~"), "BCLS", "DATA_FILE_MAP.xlsx")
DEFAULT_ROOT_MAP_PATH = os.path.join(ROOT, "DATA_FILE_MAP.xlsx")
EXCEL_MAP_PATH = os.environ.get("BCLS_DATA_MAP_XLSX")
if not EXCEL_MAP_PATH:
    EXCEL_MAP_PATH = DEFAULT_ROOT_MAP_PATH if os.path.exists(DEFAULT_ROOT_MAP_PATH) else DEFAULT_HOME_MAP_PATH
EXCEL_MAP_REQUIRED = [
    {
        "key": "life_sciences_main",
        "description": "Life Sciences sector workbook",
        "required": True,
    },
    {
        "key": "look_west_media",
        "description": "Look West media coverage workbook",
        "required": True,
    },
    {
        "key": "look_west_funding",
        "description": "Look West funding/investments workbook",
        "required": True,
    },
]
EXCEL_FILE_FALLBACKS = {
    "life_sciences_main": [
        os.path.join(ROOT, "data", "sectors", "life_sciences", "Life_Sciences_light.xlsx"),
    ],
    "look_west_media": [
        os.path.join(ROOT, "data", "look_west_strategy", "look_west_media_coverage_mock.xlsx"),
        os.path.join(ROOT, "data", "look_west_strategy", "Look West media coverage - mock.xlsx"),
    ],
    "look_west_funding": [
        os.path.join(ROOT, "data", "look_west_strategy", "lw_related_funding_investments_mock.xlsx"),
        os.path.join(ROOT, "data", "look_west_strategy", "LW related funding and investments - Mock.xlsx"),
    ],
}
CLOUD_CACHE_DIR = os.path.join(os.path.expanduser("~"), "BCLS", "cache", "excel")
CLOUD_TOKEN_CACHE_PATH = os.path.join(os.path.expanduser("~"), "BCLS", "cache", "msal_token_cache.bin")
GRAPH_SCOPES = ["Files.Read", "Sites.Read.All"]

os.chdir(ROOT)


def create_http_server(handler_cls, candidates):
    last_err = None
    for port in candidates:
        try:
            httpd = socketserver.TCPServer(("", port), handler_cls)
            actual = httpd.server_address[1]
            return httpd, actual
        except OSError as e:
            last_err = e
            continue
    if last_err:
        raise last_err
    raise OSError("No available port found")


def open_browser(port):
    import time
    time.sleep(1.0)
    webbrowser.open(f"http://localhost:{port}/dashboards/bc_dashboard_hub/html/dashboard.html")


def ensure_excel_map_template(path_str):
    if load_workbook is None or Workbook is None:
        return False, "openpyxl not available"
    p = Path(path_str)
    if p.exists():
        wb = load_workbook(p)
        changed = False
        if "FILE_MAP" not in wb.sheetnames:
            ws = wb.create_sheet("FILE_MAP", 0)
            ws.append(["key", "path", "required", "description", "notes"])
            for r in EXCEL_MAP_REQUIRED:
                ws.append([r["key"], "", "TRUE" if r.get("required") else "FALSE", r.get("description", ""), "Enter local synced SharePoint path OR SharePoint URL"])
            changed = True
        if "SETTINGS" not in wb.sheetnames:
            cfg = wb.create_sheet("SETTINGS")
            cfg.append(["key", "value", "notes"])
            cfg.append(["BCLS_GRAPH_CLIENT_ID", "", "Required for authenticated SharePoint URL download"])
            cfg.append(["BCLS_GRAPH_TENANT_ID", "organizations", "Optional; organizations/common or a tenant ID"])
            changed = True
        if changed:
            wb.save(p)
            return True, "updated"
        return True, "exists"
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "FILE_MAP"
    ws.append(["key", "path", "required", "description", "notes"])
    for r in EXCEL_MAP_REQUIRED:
        ws.append([r["key"], "", "TRUE" if r.get("required") else "FALSE", r.get("description", ""), "Enter local synced SharePoint path OR SharePoint URL"])
    cfg = wb.create_sheet("SETTINGS")
    cfg.append(["key", "value", "notes"])
    cfg.append(["BCLS_GRAPH_CLIENT_ID", "", "Required for authenticated SharePoint URL download"])
    cfg.append(["BCLS_GRAPH_TENANT_ID", "organizations", "Optional; organizations/common or a tenant ID"])
    wb.save(p)
    return True, "created"


def load_excel_map(path_str):
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to read DATA_FILE_MAP.xlsx")
    p = Path(path_str)
    if not p.exists():
        return {}
    wb = load_workbook(p, data_only=True)
    ws = wb["FILE_MAP"] if "FILE_MAP" in wb.sheetnames else wb[wb.sheetnames[0]]
    mapping = {}
    # Columns: key, path, required, description, notes
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = str(row[0] or "").strip()
        if not key:
            continue
        raw_path = str(row[1] or "").strip()
        if raw_path:
            raw_path = os.path.expandvars(raw_path)
            raw_path = os.path.expanduser(raw_path)
        required = str(row[2] or "").strip().upper() in ("TRUE", "YES", "1")
        mapping[key] = {"path": raw_path, "required": required}
    return mapping


def load_excel_settings(path_str):
    if load_workbook is None:
        return {}
    p = Path(path_str)
    if not p.exists():
        return {}
    wb = load_workbook(p, data_only=True)
    if "SETTINGS" not in wb.sheetnames:
        return {}
    ws = wb["SETTINGS"]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        k = str(row[0] or "").strip()
        v = str(row[1] or "").strip()
        if k:
            out[k] = v
    return out


def load_graph_config(path_str):
    cfg_sheet = load_excel_settings(path_str)
    client_id = (
        os.environ.get("BCLS_GRAPH_CLIENT_ID")
        or cfg_sheet.get("BCLS_GRAPH_CLIENT_ID")
        or ""
    ).strip()
    tenant_id = (
        os.environ.get("BCLS_GRAPH_TENANT_ID")
        or cfg_sheet.get("BCLS_GRAPH_TENANT_ID")
        or "organizations"
    ).strip() or "organizations"
    return {"client_id": client_id, "tenant_id": tenant_id}


def validate_excel_map(path_str):
    mapping = load_excel_map(path_str)
    checks = []
    for req in EXCEL_MAP_REQUIRED:
        key = req["key"]
        row = mapping.get(key, {})
        p = row.get("path", "")
        is_url = str(p).lower().startswith(("http://", "https://"))
        fallback_path = next((fp for fp in EXCEL_FILE_FALLBACKS.get(key, []) if os.path.exists(fp)), "")
        exists = bool((p and is_url) or (p and (not is_url) and os.path.exists(p)) or (not p and fallback_path))
        checks.append(
            {
                "key": key,
                "description": req.get("description", ""),
                "required": bool(req.get("required", True)),
                "path": p,
                "exists": exists,
                "isUrl": is_url,
                "fallbackPath": fallback_path,
            }
        )
    missing = [c for c in checks if c["required"] and not c["exists"]]
    return {"mapPath": path_str, "checks": checks, "missingRequired": missing}


class UTF8RequestHandler(http.server.SimpleHTTPRequestHandler):
    extensions_map = {
        **http.server.SimpleHTTPRequestHandler.extensions_map,
        ".html": "text/html; charset=utf-8",
        ".css": "text/css; charset=utf-8",
        ".js": "application/javascript; charset=utf-8",
        ".json": "application/json; charset=utf-8",
    }

    def _send_json(self, status, payload):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _excel_candidates_from_url(self, url):
        u = str(url or "").strip()
        if not u:
            return []
        cands = [u]
        lower = u.lower()
        if "sharepoint.com" in lower:
            # SharePoint web links often need explicit download hint.
            parts = urllib.parse.urlsplit(u)
            q = urllib.parse.parse_qs(parts.query, keep_blank_values=True)
            if "download" not in {k.lower() for k in q.keys()}:
                q["download"] = ["1"]
            new_q = urllib.parse.urlencode(q, doseq=True)
            cands.append(urllib.parse.urlunsplit((parts.scheme, parts.netloc, parts.path, new_q, parts.fragment)))
        # unique preserving order
        out = []
        seen = set()
        for x in cands:
            if x not in seen:
                out.append(x)
                seen.add(x)
        return out

    def _graph_token_cache(self):
        if msal is None:
            raise RuntimeError("Microsoft Graph auth requires 'msal'. Install with: python -m pip install msal")
        cache = msal.SerializableTokenCache()
        if os.path.exists(CLOUD_TOKEN_CACHE_PATH):
            try:
                with open(CLOUD_TOKEN_CACHE_PATH, "r", encoding="utf-8") as f:
                    cache.deserialize(f.read())
            except Exception:
                pass
        return cache

    def _save_graph_token_cache(self, cache):
        try:
            if cache and cache.has_state_changed:
                os.makedirs(os.path.dirname(CLOUD_TOKEN_CACHE_PATH), exist_ok=True)
                with open(CLOUD_TOKEN_CACHE_PATH, "w", encoding="utf-8") as f:
                    f.write(cache.serialize())
        except Exception:
            pass

    def _get_graph_access_token(self):
        cfg = load_graph_config(EXCEL_MAP_PATH)
        client_id = cfg.get("client_id", "").strip()
        tenant_id = cfg.get("tenant_id", "organizations").strip() or "organizations"
        if not client_id:
            raise RuntimeError(
                "SharePoint URL requires Microsoft Graph app registration. "
                "Set BCLS_GRAPH_CLIENT_ID in DATA_FILE_MAP.xlsx (SETTINGS sheet) or environment."
            )
        cache = self._graph_token_cache()
        app = msal.PublicClientApplication(
            client_id=client_id,
            authority=f"https://login.microsoftonline.com/{tenant_id}",
            token_cache=cache,
        )
        accounts = app.get_accounts()
        result = app.acquire_token_silent(GRAPH_SCOPES, account=accounts[0] if accounts else None)
        if not result or "access_token" not in result:
            flow = app.initiate_device_flow(scopes=GRAPH_SCOPES)
            if "user_code" not in flow:
                msg = flow.get("error_description") or flow.get("error") or json.dumps(flow, ensure_ascii=False)
                raise RuntimeError(f"Failed to start Microsoft login flow for SharePoint download. {msg}")
            print("\n[AUTH] SharePoint sign-in required. Follow the device login prompt below:")
            print(flow.get("message", "Open https://microsoft.com/devicelogin and enter the shown code."))
            result = app.acquire_token_by_device_flow(flow)
        self._save_graph_token_cache(cache)
        if not result or "access_token" not in result:
            err = (result or {}).get("error_description") or (result or {}).get("error") or "Unknown auth error"
            raise RuntimeError(f"Microsoft login failed: {err}")
        return result["access_token"]

    def _download_excel_via_graph_share(self, share_url):
        token = self._get_graph_access_token()
        share_token = "u!" + base64.urlsafe_b64encode(share_url.encode("utf-8")).decode("ascii").rstrip("=")
        graph_url = f"https://graph.microsoft.com/v1.0/shares/{share_token}/driveItem/content"
        req = urllib.request.Request(
            url=graph_url,
            method="GET",
            headers={
                "Authorization": f"Bearer {token}",
                "User-Agent": self._BROWSER_UA,
                "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*",
            },
        )
        with self._urlopen_with_fallback(req, timeout=90) as resp:
            raw = resp.read()
            ctype = str(resp.headers.get("Content-Type", "")).lower()
        if raw[:2] != b"PK":
            sample = raw[:140].decode("utf-8", errors="ignore")
            raise RuntimeError(
                f"Graph response was not an .xlsx file (content-type={ctype or 'unknown'}). Sample: {sample!r}"
            )
        return raw, graph_url

    def _cloud_cache_path(self, key, url):
        digest = hashlib.sha1(str(url).encode("utf-8")).hexdigest()[:12]
        name = f"{key}_{digest}.xlsx"
        return os.path.join(CLOUD_CACHE_DIR, name)

    def _write_cloud_cache(self, path, data):
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "wb") as f:
            f.write(data)

    def _download_excel_from_url(self, url):
        last_err = None
        for cand in self._excel_candidates_from_url(url):
            req = urllib.request.Request(
                url=cand,
                method="GET",
                headers={
                    "User-Agent": self._BROWSER_UA,
                    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*",
                    "Accept-Language": "en-CA,en;q=0.9",
                },
            )
            try:
                with self._urlopen_with_fallback(req, timeout=60) as resp:
                    raw = resp.read()
                    ctype = str(resp.headers.get("Content-Type", "")).lower()
                # XLSX is a zip archive; first bytes are PK.
                if raw[:2] != b"PK":
                    sample = raw[:140].decode("utf-8", errors="ignore")
                    raise RuntimeError(
                        f"Response was not an .xlsx file (content-type={ctype or 'unknown'}). "
                        f"Sample: {sample!r}"
                    )
                return raw, cand
            except Exception as e:
                last_err = e
        # Fallback: authenticated Microsoft Graph download for SharePoint links.
        if "sharepoint.com" in str(url).lower():
            try:
                return self._download_excel_via_graph_share(url)
            except Exception as e:
                last_err = e
        raise RuntimeError(f"Could not download Excel from URL. Last error: {last_err}")

    def end_headers(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, Accept")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(204)
        self.end_headers()

    def _read_json_body(self):
        n = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(n) if n > 0 else b"{}"
        try:
            return json.loads(raw.decode("utf-8"))
        except Exception:
            return None

    def _json_status_payload(self):
        try:
            return validate_excel_map(EXCEL_MAP_PATH)
        except Exception as e:
            return {"mapPath": EXCEL_MAP_PATH, "checks": [], "missingRequired": [], "error": str(e)}

    def _serve_excel_file_by_key(self, key):
        if not key:
            self._send_json(400, {"error": "Missing key"})
            return
        try:
            mapping = load_excel_map(EXCEL_MAP_PATH)
            row = mapping.get(key)
            if not row:
                self._send_json(404, {"error": f"Map key not found: {key}", "mapPath": EXCEL_MAP_PATH})
                return
            file_path = row.get("path", "")
            used_fallback = False
            if not file_path:
                fallback = next((fp for fp in EXCEL_FILE_FALLBACKS.get(key, []) if os.path.exists(fp)), "")
                if fallback:
                    file_path = fallback
                    used_fallback = True
                    print(f"[INFO] Using mock fallback for key '{key}': {file_path}")
                else:
                    self._send_json(404, {"error": f"No path configured for key: {key}", "mapPath": EXCEL_MAP_PATH})
                    return
            if str(file_path).lower().startswith(("http://", "https://")):
                cache_path = self._cloud_cache_path(key, file_path)
                try:
                    data, used_url = self._download_excel_from_url(file_path)
                    self._write_cloud_cache(cache_path, data)
                    src = "mapped-url"
                except Exception as download_err:
                    if os.path.exists(cache_path):
                        with open(cache_path, "rb") as f:
                            data = f.read()
                        used_url = file_path
                        src = "cloud-cache"
                        print(f"[WARN] Cloud download failed for '{key}', using cached local copy: {download_err}")
                    else:
                        raise
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(data)))
                self.send_header("X-BCLS-Data-Source", src)
                self.send_header("X-BCLS-Data-URL", used_url)
                self.end_headers()
                self.wfile.write(data)
                return

            if not os.path.exists(file_path):
                self._send_json(404, {"error": f"Configured path not found for key: {key}", "path": file_path, "mapPath": EXCEL_MAP_PATH})
                return
            with open(file_path, "rb") as f:
                data = f.read()
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Cache-Control", "no-store")
            self.send_header("Content-Length", str(len(data)))
            self.send_header("X-BCLS-Data-Source", "mock-fallback" if used_fallback else "mapped-path")
            self.end_headers()
            self.wfile.write(data)
        except Exception as e:
            self._send_json(500, {"error": "Failed to read mapped Excel file", "details": str(e), "key": key, "mapPath": EXCEL_MAP_PATH})

    def _urlopen_no_proxy(self, req, timeout):
        opener = urllib.request.build_opener(
            urllib.request.ProxyHandler({}),
            urllib.request.HTTPSHandler(context=ssl.create_default_context()),
        )
        try:
            return opener.open(req, timeout=timeout)
        except Exception:
            opener_insecure = urllib.request.build_opener(
                urllib.request.ProxyHandler({}),
                urllib.request.HTTPSHandler(context=ssl._create_unverified_context()),
            )
            return opener_insecure.open(req, timeout=timeout)

    def _urlopen_with_fallback(self, req, timeout):
        # 1) System/network proxy settings (works in managed enterprise networks)
        try:
            return urllib.request.urlopen(req, timeout=timeout)
        except Exception:
            # 2) Explicit no-proxy path (works where proxy env causes issues)
            return self._urlopen_no_proxy(req, timeout=timeout)

    def _curl_post(self, url, headers, data, timeout_sec, no_proxy=False):
        cmd = ["curl.exe", "-sS", "-L", "--max-time", str(int(timeout_sec)), "-X", "POST", url]
        for k, v in (headers or {}).items():
            cmd.extend(["-H", f"{k}: {v}"])
        cmd.extend(["--data-binary", "@-"])
        env = os.environ.copy()
        if no_proxy:
            for k in ("http_proxy", "https_proxy", "HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY", "all_proxy", "NO_PROXY", "no_proxy"):
                env[k] = ""
        proc = subprocess.run(cmd, input=(data or b""), capture_output=True, env=env)
        if proc.returncode != 0:
            err = (proc.stderr or b"").decode("utf-8", errors="replace").strip()
            raise RuntimeError(f"curl failed ({proc.returncode}): {err[:300]}")
        return proc.stdout

    _BROWSER_UA  = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
    _STC_ORIGIN  = "https://www150.statcan.gc.ca"
    _STC_REFERER = "https://www150.statcan.gc.ca/n1/en/type/data"

    def _proxy_statcan_wds(self, body):
        path = str(body.get("path", "")).strip().lstrip("/")
        payload = body.get("payload", [])
        if not path:
            self._send_json(400, {"error": "Missing path"})
            return
        if not isinstance(payload, list):
            self._send_json(400, {"error": "payload must be an array"})
            return

        url = f"https://www150.statcan.gc.ca/t1/wds/rest/{path}"
        data = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(
            url=url, data=data, method="POST",
            headers={
                "Content-Type": "application/json",
                "Accept": "application/json",
                "User-Agent": self._BROWSER_UA,
                "Origin": self._STC_ORIGIN,
                "Referer": self._STC_REFERER,
                "Accept-Language": "en-CA,en;q=0.9",
            },
        )
        try:
            with self._urlopen_with_fallback(req, timeout=45) as resp:
                raw = resp.read()
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Cache-Control", "no-store")
            self.send_header("Content-Length", str(len(raw)))
            self.end_headers()
            self.wfile.write(raw)
        except urllib.error.HTTPError as e:
            msg = e.read().decode("utf-8", errors="replace")
            print(f"[WDS proxy] HTTP {e.code}: {msg[:200]}")
            self._send_json(502, {"error": f"WDS HTTP {e.code}", "details": msg[:400]})
        except Exception as e:
            try:
                raw = self._curl_post(
                    url,
                    {
                        "Content-Type": "application/json",
                        "Accept": "application/json",
                        "User-Agent": self._BROWSER_UA,
                        "Origin": self._STC_ORIGIN,
                        "Referer": self._STC_REFERER,
                        "Accept-Language": "en-CA,en;q=0.9",
                    },
                    data,
                    45,
                    no_proxy=False,
                )
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(raw)))
                self.end_headers()
                self.wfile.write(raw)
                return
            except Exception:
                try:
                    raw = self._curl_post(
                        url,
                        {
                            "Content-Type": "application/json",
                            "Accept": "application/json",
                            "User-Agent": self._BROWSER_UA,
                            "Origin": self._STC_ORIGIN,
                            "Referer": self._STC_REFERER,
                            "Accept-Language": "en-CA,en;q=0.9",
                        },
                        data,
                        45,
                        no_proxy=True,
                    )
                    self.send_response(200)
                    self.send_header("Content-Type", "application/json; charset=utf-8")
                    self.send_header("Cache-Control", "no-store")
                    self.send_header("Content-Length", str(len(raw)))
                    self.end_headers()
                    self.wfile.write(raw)
                    return
                except Exception as e2:
                    detail = f"{type(e).__name__}: {repr(e)} | curl fallback: {type(e2).__name__}: {repr(e2)}"
                    print(f"[WDS proxy] error: {detail}")
                    self._send_json(502, {"error": "WDS proxy failed", "details": detail})

    def _proxy_statcan_csv(self, body):
        pid = str(body.get("pid", "")).strip()
        try:
            latest_n = int(float(body.get("latestN", 40) or 40))
        except Exception:
            latest_n = 40
        selected_members = body.get("selectedMembers", [])
        checked_levels = str(body.get("checkedLevels", "")).strip()
        start_date = str(body.get("startDate", "")).strip()
        end_date = str(body.get("endDate", "")).strip()
        if not pid:
            self._send_json(400, {"error": "Missing pid"})
            return
        if not isinstance(selected_members, list):
            self._send_json(400, {"error": "selectedMembers must be an array"})
            return

        q = urllib.parse.urlencode(
            {
                "pid": pid,
                "latestN": str(latest_n),
                "startDate": start_date,
                "endDate": end_date,
                "csvLocale": "en",
                "selectedMembers": json.dumps(selected_members),
                "checkedLevels": checked_levels,
            }
        )
        url = (
            "https://www150.statcan.gc.ca/t1/tbl1/en/"
            f"dtl!downloadDbLoadingData-nonTraduit.action?{q}"
        )
        req = urllib.request.Request(
            url=url, data=b" ", method="POST",
            headers={
                "Content-Type": "application/x-www-form-urlencoded",
                "Accept": "text/csv,text/plain,*/*",
                "User-Agent": self._BROWSER_UA,
                "Origin": self._STC_ORIGIN,
                "Referer": f"https://www150.statcan.gc.ca/t1/tbl1/en/tv.action?pid={pid}",
                "Accept-Language": "en-CA,en;q=0.9",
                "Accept-Encoding": "gzip, deflate, br",
                "X-Requested-With": "XMLHttpRequest",
            },
        )
        try:
            with self._urlopen_with_fallback(req, timeout=90) as resp:
                raw = resp.read()
            content = raw
            if resp.headers.get("Content-Encoding") == "gzip" or raw[:2] == b"\x1f\x8b":
                import gzip
                try:
                    content = gzip.decompress(raw)
                except Exception:
                    pass
            decoded = content.decode("utf-8-sig", errors="replace")
            if "REF_DATE" not in decoded:
                print(f"[CSV proxy] non-CSV response for pid={pid}: {decoded[:200]}")
                self._send_json(502, {
                    "error": "StatsCan did not return CSV data",
                    "details": decoded[:400],
                })
                return
            out = decoded.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/csv; charset=utf-8")
            self.send_header("Cache-Control", "no-store")
            self.send_header("Content-Length", str(len(out)))
            self.end_headers()
            self.wfile.write(out)
        except urllib.error.HTTPError as e:
            msg = e.read().decode("utf-8", errors="replace")
            print(f"[CSV proxy] HTTP {e.code}: {msg[:200]}")
            self._send_json(502, {"error": f"CSV HTTP {e.code}", "details": msg[:400]})
        except Exception as e:
            try:
                raw = self._curl_post(
                    url,
                    {
                        "Content-Type": "application/x-www-form-urlencoded",
                        "Accept": "text/csv,text/plain,*/*",
                        "User-Agent": self._BROWSER_UA,
                        "Origin": self._STC_ORIGIN,
                        "Referer": f"https://www150.statcan.gc.ca/t1/tbl1/en/tv.action?pid={pid}",
                        "Accept-Language": "en-CA,en;q=0.9",
                        "Accept-Encoding": "gzip, deflate, br",
                        "X-Requested-With": "XMLHttpRequest",
                    },
                    b" ",
                    90,
                    no_proxy=False,
                )
                decoded = raw.decode("utf-8-sig", errors="replace")
                if "REF_DATE" not in decoded:
                    self._send_json(502, {
                        "error": "StatsCan did not return CSV data",
                        "details": decoded[:400],
                    })
                    return
                out = decoded.encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "text/csv; charset=utf-8")
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(out)))
                self.end_headers()
                self.wfile.write(out)
                return
            except Exception:
                try:
                    raw = self._curl_post(
                        url,
                        {
                            "Content-Type": "application/x-www-form-urlencoded",
                            "Accept": "text/csv,text/plain,*/*",
                            "User-Agent": self._BROWSER_UA,
                            "Origin": self._STC_ORIGIN,
                            "Referer": f"https://www150.statcan.gc.ca/t1/tbl1/en/tv.action?pid={pid}",
                            "Accept-Language": "en-CA,en;q=0.9",
                            "Accept-Encoding": "gzip, deflate, br",
                            "X-Requested-With": "XMLHttpRequest",
                        },
                        b" ",
                        90,
                        no_proxy=True,
                    )
                    decoded = raw.decode("utf-8-sig", errors="replace")
                    if "REF_DATE" not in decoded:
                        self._send_json(502, {
                            "error": "StatsCan did not return CSV data",
                            "details": decoded[:400],
                        })
                        return
                    out = decoded.encode("utf-8")
                    self.send_response(200)
                    self.send_header("Content-Type", "text/csv; charset=utf-8")
                    self.send_header("Cache-Control", "no-store")
                    self.send_header("Content-Length", str(len(out)))
                    self.end_headers()
                    self.wfile.write(out)
                    return
                except Exception as e2:
                    detail = f"{type(e).__name__}: {repr(e)} | curl fallback: {type(e2).__name__}: {repr(e2)}"
                    print(f"[CSV proxy] error: {detail}")
                    self._send_json(502, {"error": "CSV proxy failed", "details": detail})

    def do_POST(self):
        try:
            if self.path == "/api/statcan-wds":
                body = self._read_json_body()
                if body is None:
                    self._send_json(400, {"error": "Invalid JSON"})
                    return
                self._proxy_statcan_wds(body)
                return

            if self.path == "/api/statcan-csv":
                body = self._read_json_body()
                if body is None:
                    self._send_json(400, {"error": "Invalid JSON"})
                    return
                self._proxy_statcan_csv(body)
                return

            super().do_POST()
        except Exception as e:
            self._send_json(500, {"error": "Local proxy handler failure", "details": str(e)})

    def do_GET(self):
        try:
            parsed = urllib.parse.urlparse(self.path)
            if parsed.path == "/api/excel-map-status":
                self._send_json(200, self._json_status_payload())
                return
            if parsed.path == "/api/excel-file":
                q = urllib.parse.parse_qs(parsed.query or "")
                key = str((q.get("key") or [""])[0]).strip()
                self._serve_excel_file_by_key(key)
                return
            super().do_GET()
        except Exception as e:
            self._send_json(500, {"error": "Local proxy GET handler failure", "details": str(e)})


Handler = UTF8RequestHandler
ok, state = ensure_excel_map_template(EXCEL_MAP_PATH)
status = validate_excel_map(EXCEL_MAP_PATH)
if state == "created":
    print(f"[INFO] Created Excel map template: {EXCEL_MAP_PATH}")
elif state == "updated":
    print(f"[INFO] Updated Excel map template with missing sheet(s): {EXCEL_MAP_PATH}")
print(f"[INFO] Excel map path in use: {EXCEL_MAP_PATH}")
if status.get("missingRequired"):
    print("[WARN] Missing required mapped Excel files:")
    for m in status["missingRequired"]:
        print(f"  - {m['key']}: {m.get('description','')} (configured path: {m.get('path') or '<blank>'})")
    print(f"[INFO] Edit mapping workbook: {EXCEL_MAP_PATH}\n")

try:
    httpd, PORT = create_http_server(Handler, PREFERRED_PORTS)
except OSError as e:
    if getattr(e, "winerror", None) == 10048:
        print("[INFO] Port 8080 is already in use. Reusing existing local server on 8080.")
        open_browser(8080)
        raise SystemExit(0)
    raise

threading.Thread(target=open_browser, args=(PORT,), daemon=True).start()
with httpd:
    print(f"[OK] Serving dashboard workspace from: {ROOT}")
    print(f"[OK] Hub URL:      http://localhost:{PORT}/dashboards/bc_dashboard_hub/html/dashboard.html")
    print(f"[OK] Macro URL:    http://localhost:{PORT}/dashboards/bc_macroeconomy/html/dashboard.html")
    print(f"   Press Ctrl+C to stop.\n")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\nServer stopped.")
